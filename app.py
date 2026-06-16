"""
app.py — AI-Powered Agile Sprint Manager (Jira-style)
CVR College of Engineering | IOMP Batch 19
Run: streamlit run app.py
"""
from zoneinfo import ZoneInfo
from dotenv import load_dotenv
load_dotenv()

import streamlit as st
import pandas as pd, numpy as np, sys, os, io
from datetime import date, timedelta, datetime
import matplotlib; matplotlib.use("Agg")
import matplotlib.pyplot as plt

sys.path.insert(0, os.path.dirname(__file__))
from data.db_setup import create_tables, seed_data, get_conn, log_activity
from auth import signup_user, login_user, logout_user
from ai_module import generate_insights
from blocker_detection import detect_blockers
from models.ml_engine import (
    CompletionTimePredictor, RiskDetector, AutoBlockerDetector,
    VelocityForecaster, RecommendationEngine, compute_sprint_health
)
import time
start_time = time.time()

st.set_page_config(page_title="SprintAI", page_icon="⚡",
                   layout="wide", initial_sidebar_state="expanded")

# ── session state for active tab ──────────────────────────────────────────────
import urllib.parse
import hmac
import hashlib
import json
import base64
import time
import streamlit.components.v1 as components

SECRET_KEY = os.getenv("SESSION_SECRET", os.getenv("SUPABASE_ANON_KEY", "sprintai_default_secret_key")).encode()

def get_cookies():
    try:
        if hasattr(st, "context") and hasattr(st.context, "cookies"):
            return dict(st.context.cookies)
    except Exception:
        pass
    try:
        from streamlit.web.server.websocket_headers import _get_websocket_headers
        headers = _get_websocket_headers()
        if not headers:
            return {}
        cookie_str = headers.get("Cookie", "")
        cookies = {}
        if cookie_str:
            for cookie in cookie_str.split(";"):
                parts = cookie.split("=", 1)
                if len(parts) == 2:
                    cookies[parts[0].strip()] = urllib.parse.unquote(parts[1].strip())
        return cookies
    except Exception:
        return {}

def set_cookie(name, value, days=None):
    cookie_val = urllib.parse.quote(value)
    if days is not None:
        js = f"""
        <script>
        var date = new Date();
        date.setTime(date.getTime() + ({days}*24*60*60*1000));
        var expires = "; expires=" + date.toUTCString();
        window.parent.document.cookie = "{name}=" + "{cookie_val}" + expires + "; path=/; SameSite=Lax";
        </script>
        """
    else:
        js = f"""
        <script>
        window.parent.document.cookie = "{name}=" + "{cookie_val}" + "; path=/; SameSite=Lax";
        </script>
        """
    components.html(js, height=0, width=0)

def delete_cookie(name):
    js = f"""
    <script>
    window.parent.document.cookie = "{name}=; expires=Thu, 01 Jan 1970 00:00:00 UTC; path=/;";
    </script>
    """
    components.html(js, height=0, width=0)

def create_session_token(email, role, name):
    data = {
        "email": email,
        "role": role,
        "name": name,
        "expires": time.time() + 7 * 24 * 60 * 60  # 7 days
    }
    payload = base64.b64encode(json.dumps(data).encode()).decode()
    signature = hmac.new(SECRET_KEY, payload.encode(), hashlib.sha256).hexdigest()
    return f"{payload}.{signature}"

def verify_session_token(token):
    try:
        parts = token.split(".")
        if len(parts) != 2:
            return None
        payload, signature = parts
        expected_sig = hmac.new(SECRET_KEY, payload.encode(), hashlib.sha256).hexdigest()
        if not hmac.compare_digest(signature, expected_sig):
            return None
        data = json.loads(base64.b64decode(payload.encode()).decode())
        if time.time() > data.get("expires", 0):
            return None
        return data
    except Exception:
        return None

cookies = get_cookies()

if "tab" not in st.session_state:
    raw_tab = cookies.get("session_tab", "Dashboard")
    tab_mapping = {"Summary": "Dashboard", "List": "Projects", "Daily Progress": "Daily Updates"}
    st.session_state.tab = tab_mapping.get(raw_tab, raw_tab)

if "user" not in st.session_state or "role" not in st.session_state or "user_name" not in st.session_state:
    st.session_state["user"] = None
    st.session_state["role"] = None
    st.session_state["user_name"] = None

# Auto-login if session token is found in cookies
if not st.session_state["user"] and not st.session_state.get("logout_triggered"):
    token = cookies.get("session_token")
    if token:
        data = verify_session_token(token)
        if data:
            st.session_state["user"] = data["email"]
            st.session_state["role"] = data["role"]
            st.session_state["user_name"] = data["name"]
            st.rerun()

def render_landing_page():
    qp = st.query_params
    if "go" in qp:
        dest = qp["go"]
        st.query_params.clear()
        st.session_state["auth_page"] = dest
        st.rerun()

    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Outfit:wght@500;700;800&family=Inter:wght@300;400;500;600;700&display=swap');

    :root { color-scheme: dark !important; }
    html { scroll-behavior: smooth !important; color-scheme: dark !important; }

    @media (prefers-color-scheme: light) {
        :root, html, body, .main, [data-testid="stApp"], [data-testid="stAppViewContainer"] {
            color-scheme: dark !important;
            background-color: #0d1117 !important;
            color: #e6edf3 !important;
        }
        section.main, [data-testid="stAppViewContainer"] {
            background-color: #0d1117 !important;
        }
    }

    section.main {
        background: radial-gradient(circle at top right, rgba(99,102,241,0.15), transparent 40%),
                    radial-gradient(circle at bottom left, rgba(168,85,247,0.1), transparent 40%),
                    #0D1117 !important;
        font-family: 'Inter', sans-serif !important;
    }

    /* Hide default Streamlit top padding */
    .block-container { padding-top: 2rem !important; }

    /* ── NAVBAR ── */
    .nav-wrapper {
        display: flex;
        align-items: center;
        justify-content: space-between;
        padding: 16px 0;
        border-bottom: 1px solid rgba(255,255,255,0.06);
        margin-bottom: 0;
    }
    .nav-left {
        display: flex;
        align-items: center;
        gap: 10px;
        font-family: 'Outfit', sans-serif;
        font-size: 22px;
        font-weight: 800;
        color: #ffffff;
        letter-spacing: -0.5px;
    }
    .logo-glow {
        background: linear-gradient(135deg, #58a6ff, #a855f7);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        filter: drop-shadow(0 0 10px rgba(88,166,255,0.4));
    }
    .nav-right {
        display: flex;
        align-items: center;
        gap: 8px;
    }
    .nav-link {
        display: inline-block;
        padding: 7px 16px;
        text-decoration: none;
        color: #94A3B8;          /* muted, not bright */
        font-weight: 500;
        font-size: 13.5px;
        border-radius: 6px;
        transition: all 0.2s ease;
        border: none;            /* ← remove the box */
        background: transparent;
    }
    .nav-link.active {
        color: #58a6ff;              /* blue tint to show it's clickable */
        background: transparent;
        border: none;
    }
    .nav-link:hover {
        color: #ffffff;
        background: rgba(255, 255, 255, 0.08);  /* subtle highlight on hover */
        border-radius: 6px;
    }
    .nav-btn {
        display: inline-block;
        padding: 7px 18px;
        text-decoration: none;
        font-weight: 600;
        font-size: 13px;
        border-radius: 7px;
        transition: all 0.2s ease;
        cursor: pointer;
    }
    .nav-btn-ghost {
        color: #94A3B8;
        background: transparent;
        border: none;            /* ← no border on Login */
        font-weight: 500;
    }
    .nav-btn-ghost:hover {
        color: #ffffff;
        background: rgba(255,255,255,0.06);
        border: none;
    }
    .nav-btn-primary {
        color: #ffffff;
        background: linear-gradient(135deg, #4f46e5, #7c3aed);
        box-shadow: 0 2px 10px rgba(124, 58, 237, 0.35);
        border: none;
        border-radius: 7px;
        padding: 7px 18px;
        text-decoration: none;
    }
    .nav-btn-primary:hover {
        background: linear-gradient(135deg, #6366f1, #a855f7);
        box-shadow: 0 4px 14px rgba(168, 85, 247, 0.45);
        color: #ffffff;
        text-decoration: none;
    }
    /* Pull buttons up into navbar */
    /* Align Features link vertically in column */
    [data-testid="stMarkdownContainer"] a.nav-link {
        display: flex !important;
        align-items: center !important;
        height: 38px !important;
        margin-top: 4px !important;
    }
    [data-testid="stHorizontalBlock"] {
        margin-top: -58px !important;
        margin-bottom: 0 !important;
    }
    /* All nav buttons ghost by default */
    [data-testid="stButton"] button {
        background: transparent !important;
        border: none !important;
        color: #94A3B8 !important;
        font-size: 13px !important;
        font-weight: 500 !important;
    }
    [data-testid="stButton"] button:hover {
        color: #ffffff !important;
        background: rgba(255,255,255,0.06) !important;
    }

    /* Sign Up only gets purple */
    [data-testid="stButton"] button[kind="primary"] {
        background: linear-gradient(135deg, #4f46e5, #7c3aed) !important;
        border: none !important;
        color: #ffffff !important;
        font-weight: 600 !important;
    }
    [data-testid="stButton"] button[kind="primary"]:hover {
        background: linear-gradient(135deg, #6366f1, #a855f7) !important;
    }
    /* Remove Streamlit default anchor styling */
    [data-testid="stMarkdownContainer"] a {
        color: #94A3B8 !important;
        background: transparent !important;
        text-decoration: none !important;
        border: none !important;
        outline: none !important;
    }
    [data-testid="stMarkdownContainer"] a:hover {
        color: #ffffff !important;
        background: rgba(255,255,255,0.06) !important;
        border-radius: 6px !important;
    }
    a.nav-features-link,
    a.nav-features-link:link,
    a.nav-features-link:visited,
    a.nav-features-link:active {
        display: flex !important;
        align-items: center !important;
        justify-content: center !important;
        height: 38px !important;
        margin-top: 5px !important;
        color: #94A3B8 !important;
        font-size: 13px !important;
        font-weight: 500 !important;
        text-decoration: none !important;
        background: transparent !important;
        border: none !important;
        outline: none !important;
        box-shadow: none !important;
        border-radius: 6px !important;
        -webkit-text-fill-color: #94A3B8 !important;
    }
    a.nav-features-link:hover {
        color: #ffffff !important;
        -webkit-text-fill-color: #ffffff !important;
        background: rgba(255,255,255,0.06) !important;
    }
    /* ── HERO ── */
    .hero-sec {
        text-align: center;
        padding: 60px 0 60px 0;
        max-width: 850px;
        margin: 0 auto;
    }
    .hero-tagline {
        background: linear-gradient(135deg, #58a6ff 0%, #7c3aed 50%, #a855f7 100%);
        -webkit-background-clip: text;
        -webkit-text-fill-color: transparent;
        font-family: 'Outfit', sans-serif;
        font-size: 58px;
        font-weight: 800;
        line-height: 1.15;
        letter-spacing: -1.5px;
        margin-bottom: 24px;
    }
    .hero-subtext {
        font-size: 17px;
        color: #CBD5E1;
        line-height: 1.6;
        margin-bottom: 36px;
        font-weight: 400;
    }

    /* ── FEATURES ── */
    .section-title {
        text-align: center;
        font-family: 'Outfit', sans-serif;
        font-size: 32px;
        font-weight: 700;
        color: #F8FAFC;
        margin-top: 80px;
        margin-bottom: 10px;
        letter-spacing: -0.5px;
    }
    .section-subtitle {
        text-align: center;
        font-size: 14px;
        color: #94A3B8;
        margin-bottom: 44px;
    }
    .feat-grid {
        display: grid;
        grid-template-columns: repeat(4, 1fr);
        gap: 20px;
        margin-bottom: 60px;
    }
    @media (max-width: 1024px) { .feat-grid { grid-template-columns: repeat(2, 1fr); } }
    @media (max-width: 600px)  { .feat-grid { grid-template-columns: 1fr; } }
    .feat-card {
        background: rgba(22,27,34,0.45);
        border: 1px solid rgba(255,255,255,0.05);
        border-radius: 12px;
        padding: 24px;
        transition: all 0.3s ease;
    }
    .feat-card:hover {
        transform: translateY(-4px);
        background: rgba(22,27,34,0.7);
        border-color: #58a6ff !important;
        box-shadow: 0 0 15px rgba(88,166,255,0.25);
    }
    
    /* Bright, premium button styles for landing page only */
    div.stButton > button[kind="primary"],
    div.stButton > button[data-testid="baseButton-primary"] {
        background: linear-gradient(135deg, #10b981, #059669) !important;
        border: none !important;
        box-shadow: 0 0 10px rgba(16, 185, 129, 0.4) !important;
        transition: all 0.3s ease !important;
    }
    div.stButton > button[kind="primary"] *,
    div.stButton > button[data-testid="baseButton-primary"] * {
        color: #ffffff !important;
        font-weight: 700 !important;
        font-size: 14px !important;
    }
    div.stButton > button[kind="primary"]:hover,
    div.stButton > button[data-testid="baseButton-primary"]:hover {
        transform: scale(1.05) !important;
        box-shadow: 0 0 15px rgba(16, 185, 129, 0.6) !important;
    }
    div.stButton > button[kind="secondary"],
    div.stButton > button[data-testid="baseButton-secondary"] {
        background: rgba(88, 166, 255, 0.15) !important;
        border: 1px solid #58a6ff !important;
        box-shadow: 0 0 10px rgba(88, 166, 255, 0.2) !important;
        transition: all 0.3s ease !important;
    }
    div.stButton > button[kind="secondary"] *,
    div.stButton > button[data-testid="baseButton-secondary"] * {
        color: #ffffff !important;
        font-weight: 700 !important;
        font-size: 14px !important;
    }
    div.stButton > button[kind="secondary"]:hover,
    div.stButton > button[data-testid="baseButton-secondary"]:hover {
        background: rgba(88, 166, 255, 0.25) !important;
        transform: scale(1.05) !important;
        box-shadow: 0 0 15px rgba(88, 166, 255, 0.4) !important;
    }
    
    /* Global text color override for all landing page buttons */
    div.stButton > button,
    div.stButton > button * {
        color: #ffffff !important;
        font-weight: 700 !important;
    }
    </style>
    """, unsafe_allow_html=True)

    # ── NAVBAR — 100% HTML, no Streamlit columns ──
    st.markdown("""
    <div class="nav-wrapper">
    <div class="nav-left">
        <span>⚡</span><span class="logo-glow">SprintAI</span>
    </div>
    </div>
    """, unsafe_allow_html=True)
    nc1, nc2, nc3, nc4, nc5 = st.columns([5.5, 0.6, 0.7, 0.6, 0.8])
    # with nc2:
    #     st.markdown("""
    #         <a href="#features" class="nav-features-link">Features</a>
    #     """, unsafe_allow_html=True)
    with nc3:
        
        st.markdown("""
        <a href="#features" style="
            display: inline-flex;
            align-items: center;
            justify-content: center;
            text-decoration: none;
            background: linear-gradient(135deg, #7c3aed, #a855f7);
            border: none;
            color: #ffffff;
            font-weight: 700;
            font-size: 14px;
            padding: 8px 16px;
            border-radius: 8px;
            width: 100%;
            box-sizing: border-box;
            text-align: center;
            transition: all 0.3s ease;
            height: 38px;
            margin-top: 4px;
            box-shadow: 0 0 10px rgba(124,58,237,.4);
        "
        onmouseover="this.style.transform='scale(1.05)';
                    this.style.boxShadow='0 0 15px rgba(124,58,237,.6)'"
        onmouseout="this.style.transform='scale(1)';
                    this.style.boxShadow='0 0 10px rgba(124,58,237,.4)'">
            Features
        </a>
        """, unsafe_allow_html=True)
    with nc4:
        if st.button("Login", key="nav_login", use_container_width=True):
            st.session_state["auth_page"] = "login"
            st.rerun()
    with nc5:
        if st.button("Sign Up", key="nav_signup", use_container_width=True, type="primary"):
            st.session_state["auth_page"] = "signup"
            st.rerun()
    # ── HERO ──
    st.markdown("""
    <div class="hero-sec">
      <div class="hero-tagline">Transform Your Sprints With AI-Powered Intelligence</div>
      <div class="hero-subtext">Plan smarter, predict risks, track team performance, identify blockers,
        and deliver successful sprints with AI assistance.</div>
    </div>
    """, unsafe_allow_html=True)

    # ── FEATURES ──
    st.markdown('<div id="features" class="section-title">SprintAI Platform Features</div>', unsafe_allow_html=True)
    st.markdown('<div class="section-subtitle">A comprehensive suite of tools built specifically for agile development team collaboration.</div>', unsafe_allow_html=True)

    features = [
        ("🤖", "AI Sprint Planning",      "Automatically analyze sprint workload and planning."),
        ("📋", "Task Management",         "Create, assign and track tasks efficiently."),
        ("⚠️", "Risk Prediction",          "Identify blockers and sprint risks early."),
        ("📝", "Daily Updates",           "Members submit daily work logs and blockers."),
        ("👥", "Team Management",         "Manage team members and assignments."),
        ("📊", "Sprint Analytics",        "Track velocity, completion rates and progress."),
        ("🧠", "AI Insights",             "Receive intelligent recommendations and health analysis."),
        ("📄", "PDF Reports",             "Generate detailed member, sprint and team reports."),
        ("🚫", "Auto Blocker Detection",  "Automatically identify stalled tasks."),
        ("📈", "Velocity Tracking",       "Monitor planned vs actual sprint performance."),
        ("🎯", "Sprint Health Grading",   "AI-generated sprint health scores and grades."),
        ("📦", "Kanban Board",            "Visual task management across sprint stages."),
    ]

    cards_html = '<div class="feat-grid">'
    for icon, title, desc in features:
        cards_html += f'<div class="feat-card"><div class="feat-icon">{icon}</div><div class="feat-title">{title}</div><div class="feat-desc">{desc}</div></div>'
    cards_html += '</div>'
    st.markdown(cards_html, unsafe_allow_html=True)

    st.markdown("""
    <div style="padding:40px 0;text-align:center;border-top:1px solid rgba(255,255,255,0.05);
                margin-top:80px;color:#8b949e;font-size:12px;">
        <div>⚡ SprintAI Agile Manager &amp; Analytics Suite · CVR College of Engineering</div>
        <div style="margin-top:8px;font-size:11px;color:#58a6ff;">Built for Agile Teams with Love.</div>
    </div>
    """, unsafe_allow_html=True)

def render_login_page():
    st.markdown("""
    <style>
    section.main { background: radial-gradient(circle at 30% 20%, rgba(99, 102, 241, 0.12) 0%, #0d1117 100%) !important; }
    
    /* Highlight stForm border */
    div[data-testid="stForm"] {
        border: 1.5px solid rgba(88, 166, 255, 0.4) !important;
        box-shadow: 0 0 15px rgba(88, 166, 255, 0.15) !important;
        border-radius: 12px !important;
        background: rgba(22, 27, 34, 0.45) !important;
        margin-top: -1px !important;
    }
    
    /* Make submit buttons glow */
    div.stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #ef4444, #f97316) !important;
        color: #ffffff !important;
        border: none !important;
        font-weight: 700 !important;
        box-shadow: 0 0 12px rgba(239, 68, 68, 0.45) !important;
        transition: all 0.3s ease !important;
    }
    div.stButton > button[kind="primary"] * {
        color: #ffffff !important;
        font-weight: 700 !important;
    }
    div.stButton > button[kind="primary"]:hover {
        transform: scale(1.02) !important;
        box-shadow: 0 0 18px rgba(239, 68, 68, 0.65) !important;
    }
    
    /* Make back buttons and other secondary buttons brightly glowing */
    div.stButton > button:not([kind="primary"]) {
        background: rgba(88, 166, 255, 0.15) !important;
        color: #ffffff !important;
        border: 1px solid #58a6ff !important;
        font-weight: 700 !important;
        box-shadow: 0 0 10px rgba(88, 166, 255, 0.2) !important;
        transition: all 0.3s ease !important;
    }
    div.stButton > button:not([kind="primary"]) * {
        color: #ffffff !important;
        font-weight: 700 !important;
    }
    div.stButton > button:not([kind="primary"]):hover {
        background: rgba(88, 166, 255, 0.25) !important;
        transform: scale(1.02) !important;
        box-shadow: 0 0 15px rgba(88, 166, 255, 0.4) !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Wrap in centered container to avoid broad UI
    _, center_col, _ = st.columns([0.5, 2.0, 0.5])
    with center_col:
        if st.button("← Back to Home", key="back_to_home_login"):
            st.session_state["auth_page"] = "landing"
            st.rerun()
            
        st.markdown("<br><br>", unsafe_allow_html=True)
        
        col1, col2 = st.columns([1, 1.1])
        with col1:
            st.markdown('<div style="padding: 30px; background: rgba(30, 27, 75, 0.25); border: 1.5px solid rgba(99, 102, 241, 0.45); box-shadow: 0 0 15px rgba(99, 102, 241, 0.2); border-radius: 12px; height: 100%;"><div style="font-family: \'Outfit\', sans-serif; font-size: 32px; font-weight: 800; color: #ffffff; margin-bottom: 12px;">⚡ SprintAI</div><div style="color: #94a3b8; font-size: 14px; margin-bottom: 24px; line-height: 1.5;">Access your internal team workspace to participate in active sprint cycles and log daily metrics.</div><div style="margin-top: 20px;"><div style="display:flex; gap:12px; margin-bottom:20px;"><div style="font-size: 20px;">🤖</div><div><div style="font-weight: 600; color: #ffffff; font-size: 13.5px;">AI-Driven Assistant</div><div style="color: #94a3b8; font-size: 11.5px; margin-top: 2px;">Assists scrum masters with analytics, forecasts, and risks.</div></div></div><div style="display:flex; gap:12px; margin-bottom:20px;"><div style="font-size: 20px;">📋</div><div><div style="font-weight: 600; color: #ffffff; font-size: 13.5px;">Team Workspace</div><div style="color: #94a3b8; font-size: 11.5px; margin-top: 2px;">Visual boards, timeline updates, and progress reporting.</div></div></div></div></div>', unsafe_allow_html=True)
            
        with col2:
            st.markdown('<div style="padding: 24px; background: rgba(22, 27, 34, 0.4); border: 1.5px solid rgba(88, 166, 255, 0.4); border-bottom: none; box-shadow: 0 -5px 15px rgba(88, 166, 255, 0.05); border-radius: 12px; border-bottom-left-radius: 0; border-bottom-right-radius: 0;"><div style="font-family: \'Outfit\', sans-serif; font-size: 24px; font-weight: 700; color: #ffffff; margin-bottom: 4px;">Sign In</div><div style="color: #94a3b8; font-size: 12.5px;">Enter your login credentials below.</div></div>', unsafe_allow_html=True)
            with st.form("login_form_new"):
                lu = st.text_input("Work Email Address", placeholder="name@company.com")
                lp = st.text_input("Password", type="password", placeholder="Enter your password")
                st.markdown("<br>", unsafe_allow_html=True)
                if st.form_submit_button("Sign In to Space", type="primary", use_container_width=True):
                    st.session_state["logout_triggered"] = False
                    ok, msg = login_user(lu, lp)
                    if ok:
                        st.success(msg)
                        st.rerun()
                    st.error(msg)
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<div style="text-align: center; font-size: 12px; color: #94a3b8; margin-bottom: 8px;">Don\'t have an account?</div>', unsafe_allow_html=True)
            if st.button("Create a New Workspace Account", use_container_width=True):
                st.session_state["auth_page"] = "signup"
                st.rerun()


def render_signup_page():
    st.markdown("""
    <style>
    section.main { background: radial-gradient(circle at 70% 80%, rgba(124, 58, 237, 0.12) 0%, #0d1117 100%) !important; }
    
    /* Highlight stForm border */
    div[data-testid="stForm"] {
        border: 1.5px solid rgba(124, 58, 237, 0.55) !important;
        box-shadow: 0 0 15px rgba(124, 58, 237, 0.2) !important;
        border-radius: 12px !important;
        background: rgba(22, 27, 34, 0.45) !important;
        margin-top: -1px !important;
    }
    
    /* Make submit buttons glow */
    div.stButton > button[kind="primary"] {
        background: linear-gradient(135deg, #ef4444, #f97316) !important;
        color: #ffffff !important;
        border: none !important;
        font-weight: 700 !important;
        box-shadow: 0 0 12px rgba(239, 68, 68, 0.45) !important;
        transition: all 0.3s ease !important;
    }
    div.stButton > button[kind="primary"] * {
        color: #ffffff !important;
        font-weight: 700 !important;
    }
    div.stButton > button[kind="primary"]:hover {
        transform: scale(1.02) !important;
        box-shadow: 0 0 18px rgba(239, 68, 68, 0.65) !important;
    }
    
    /* Make back buttons and other secondary buttons brightly glowing */
    div.stButton > button:not([kind="primary"]) {
        background: rgba(124, 58, 237, 0.15) !important;
        color: #ffffff !important;
        border: 1px solid #7c3aed !important;
        font-weight: 700 !important;
        box-shadow: 0 0 10px rgba(124, 58, 237, 0.2) !important;
        transition: all 0.3s ease !important;
    }
    div.stButton > button:not([kind="primary"]) * {
        color: #ffffff !important;
        font-weight: 700 !important;
    }
    div.stButton > button:not([kind="primary"]):hover {
        background: rgba(124, 58, 237, 0.25) !important;
        transform: scale(1.02) !important;
        box-shadow: 0 0 15px rgba(124, 58, 237, 0.4) !important;
    }
    </style>
    """, unsafe_allow_html=True)
    
    # Wrap in centered container to avoid broad UI
    _, center_col, _ = st.columns([0.5, 2.0, 0.5])
    with center_col:
        if st.button("← Back to Home", key="back_to_home_signup"):
            st.session_state["auth_page"] = "landing"
            st.rerun()
            
        st.markdown("<br><br>", unsafe_allow_html=True)
        
        col1, col2 = st.columns([1, 1.1])
        with col1:
            st.markdown('<div style="padding: 30px; background: rgba(124, 58, 237, 0.15); border: 1.5px solid rgba(124, 58, 237, 0.55); box-shadow: 0 0 15px rgba(124, 58, 237, 0.25); border-radius: 12px; height: 100%;"><div style="font-family: \'Outfit\', sans-serif; font-size: 32px; font-weight: 800; color: #ffffff; margin-bottom: 12px;">⚡ Join SprintAI</div><div style="color: #94a3b8; font-size: 14px; margin-bottom: 24px; line-height: 1.5;">Create an account to start managing sprints, logging updates, and utilizing AI-driven task risks.</div><div style="margin-top: 20px;"><div style="display:flex; gap:12px; margin-bottom:20px;"><div style="font-size: 20px;">🔑</div><div><div style="font-weight: 600; color: #ffffff; font-size: 13.5px;">Role-Based Account Setup</div><div style="color: #94a3b8; font-size: 11.5px; margin-top: 2px;">Register as a Manager (project owner) or Member (collaborator).</div></div></div><div style="display:flex; gap:12px; margin-bottom:20px;"><div style="font-size: 20px;">📊</div><div><div style="font-weight: 600; color: #ffffff; font-size: 13.5px;">Predictive Insights Integration</div><div style="color: #94a3b8; font-size: 11.5px; margin-top: 2px;">Automatically calculates velocity average & story points.</div></div></div></div></div>', unsafe_allow_html=True)
            
        with col2:
            st.markdown('<div style="padding: 24px; background: rgba(22, 27, 34, 0.4); border: 1.5px solid rgba(124, 58, 237, 0.55); border-bottom: none; box-shadow: 0 -5px 15px rgba(124, 58, 237, 0.05); border-radius: 12px; border-bottom-left-radius: 0; border-bottom-right-radius: 0;"><div style="font-family: \'Outfit\', sans-serif; font-size: 24px; font-weight: 700; color: #ffffff; margin-bottom: 4px;">Create Account</div><div style="color: #94a3b8; font-size: 12.5px;">Fill in details below to register.</div></div>', unsafe_allow_html=True)
            with st.form("signup_form_new"):
                sfn = st.text_input("Full Name", placeholder="e.g. John Doe")
                su = st.text_input("Work Email Address", placeholder="name@company.com")
                sp = st.text_input("Create Password", type="password", placeholder="Minimum 8 characters")
                sp_conf = st.text_input("Confirm Password", type="password", placeholder="Re-enter password")
                sr = st.selectbox("Workspace Role", ["Member", "Manager"])
                st.markdown("<br>", unsafe_allow_html=True)
                if st.form_submit_button("Register Account", type="primary", use_container_width=True):
                    if sp != sp_conf:
                        st.error("Passwords do not match.")
                    elif len(sp) < 8:
                        st.error("Password must be at least 8 characters.")
                    else:
                        ok, msg = signup_user(su, sp, sr, full_name=sfn)
                        if ok:
                            st.success(msg)
                            st.session_state["auth_page"] = "login"
                            st.rerun()
                        else:
                            st.error(msg)
            st.markdown("<br>", unsafe_allow_html=True)
            st.markdown('<div style="text-align: center; font-size: 12px; color: #94a3b8; margin-bottom: 8px;">Already have an account?</div>', unsafe_allow_html=True)
            if st.button("Sign In to Existing Workspace", use_container_width=True):
                st.session_state["auth_page"] = "login"
                st.rerun()


def render_auth_gate():
    if "auth_page" not in st.session_state:
        st.session_state["auth_page"] = "landing"
        
    page = st.session_state["auth_page"]
    if page == "landing":
        render_landing_page()
    elif page == "login":
        render_login_page()
    elif page == "signup":
        render_signup_page()

# ── CSS ───────────────────────────────────────────────────────────────────────
st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&family=DM+Mono:wght@400;500&display=swap');

/* ── Force dark mode regardless of OS/browser light-theme preference ── */
:root {
    color-scheme: dark !important;
}

/* Re-apply ALL dark-theme values when the OS is set to light mode */
@media (prefers-color-scheme: light) {
    :root, html, body {
        color-scheme: dark !important;
        --background-color: #0d1117 !important;
        --secondary-background-color: #161b22 !important;
        --text-color: #e6edf3 !important;
        --primary-color: #ef4444 !important;
    }
    /* Streamlit app containers */
    [data-testid="stApp"],
    [data-testid="stAppViewContainer"],
    [data-testid="stHeader"],
    section.main,
    .main {
        background-color: #0d1117 !important;
        color: #e6edf3 !important;
    }
    /* Sidebar */
    section[data-testid="stSidebar"] {
        background-color: #161b22 !important;
        border-right: 1px solid #30363d !important;
    }
    /* Inputs */
    .stTextInput > div > div > input,
    .stTextArea > div > div > textarea,
    .stNumberInput input,
    .stSelectbox > div > div,
    .stDateInput input,
    .stMultiSelect > div {
        background-color: #21262d !important;
        border-color: #30363d !important;
        color: #e6edf3 !important;
    }
    /* Buttons */
    .stButton > button {
        background-color: #21262d !important;
        color: #c9d1d9 !important;
        border-color: #30363d !important;
    }
    .stButton > button[kind="primary"] {
        background-color: #238636 !important;
        color: #ffffff !important;
    }
    /* Forms */
    .stForm {
        background-color: #161b22 !important;
        border-color: #30363d !important;
    }
    /* Expanders */
    .streamlit-expanderHeader {
        background-color: #161b22 !important;
        border-color: #30363d !important;
    }
    .streamlit-expanderContent {
        background-color: #0d1117 !important;
        border-color: #30363d !important;
    }
    /* Text & headings */
    h1, h2, h3 { color: #e6edf3 !important; }
    p, li, .stMarkdown { color: #8b949e !important; }
    label { color: #8b949e !important; }
    /* Metrics */
    [data-testid="stMetricValue"] { color: #e6edf3 !important; }
    [data-testid="stMetricLabel"] { color: #8b949e !important; }
    /* DataFrames */
    .stDataFrame { border-color: #30363d !important; }
}

html,body,[class*="css"]  { font-family:'Inter',sans-serif; }
code,.stCode              { font-family:'DM Mono',monospace !important; }
.main                     { background:#0d1117; }
.block-container          { padding:68px 0 32px 0 !important; max-width:100% !important; }
section[data-testid="stSidebar"] { background:#161b22; border-right:1px solid #30363d; }
section[data-testid="stSide
bar"] .block-container { padding:16px !important; }

/* hide only deploy button, keep sidebar toggle visible */
#MainMenu { visibility:hidden; }
footer    { visibility:hidden; }

/* headings & text */
h1,h2,h3              { color:#e6edf3 !important; }
p,li,.stMarkdown      { color:#8b949e; font-size:13px; }
label                 { color:#8b949e !important; font-size:12px !important; }
hr                    { border-color:#21262d !important; margin:10px 0 !important; }

/* inputs */
.stTextInput>div>div>input,
.stTextArea>div>div>textarea,
.stNumberInput input,
.stSelectbox>div>div  { background:#21262d !important; border:1px solid #30363d !important;
                         color:#e6edf3 !important; border-radius:6px !important; font-size:13px !important; }
.stDateInput input    { background:#21262d !important; color:#e6edf3 !important; border-radius:6px !important; }
.stMultiSelect>div    { background:#21262d !important; border:1px solid #30363d !important; border-radius:6px !important; }

/* buttons */
.stButton>button      { border-radius:6px !important; font-weight:500 !important;
                         font-size:12px !important; border:none !important; }
.stButton>button[kind="primary"]   { background:#238636 !important; color:#fff !important; }
.stButton>button[kind="secondary"] { background:#21262d !important; color:#c9d1d9 !important; border:1px solid #30363d !important; }

/* form */
.stForm               { background:#161b22 !important; border:1px solid #30363d !important;
                         border-radius:8px !important; padding:16px !important; }

/* metric */
[data-testid="stMetricValue"] { color:#e6edf3 !important; font-weight:700 !important; font-size:1.4rem !important; }
[data-testid="stMetricLabel"] { color:#8b949e !important; font-size:11px !important; text-transform:uppercase; letter-spacing:.5px; }

/* dataframe */
.stDataFrame          { border-radius:8px; overflow:hidden; border:1px solid #30363d !important; }

/* expander */
.streamlit-expanderHeader   { background:#161b22 !important; border-radius:6px !important; border:1px solid #30363d !important; }
.streamlit-expanderContent  { background:#0d1117 !important; border:1px solid #30363d !important; border-top:none !important; }

/* ── TOP NAV BAR ── */
.top-nav              { background:#161b22; border-bottom:1px solid #30363d;
                         padding:0; display:flex; align-items:stretch; width:100%; }
.top-nav-logo         { display:flex; align-items:center; gap:8px; padding:0 20px;
                         border-right:1px solid #30363d; min-width:200px; }
.proj-badge           { width:28px; height:28px; border-radius:6px;
                         background:linear-gradient(135deg,#f97316,#ef4444);
                         display:flex; align-items:center; justify-content:center; font-size:14px; }
.proj-name            { color:#e6edf3; font-weight:700; font-size:14px; }

/* TAB BUTTONS — these override Streamlit button styling */
div[data-testid="column"] .stButton>button.tab-btn {
    background: transparent !important;
    color: #8b949e !important;
    border: none !important;
    border-bottom: 2px solid transparent !important;
    border-radius: 0 !important;
    padding: 14px 14px 12px !important;
    font-size: 13px !important;
    font-weight: 500 !important;
    height: auto !important;
    width: 100% !important;
}

/* Page content */
.page-wrap            { padding:20px 24px; }

/* cards */
.card                 { background:#161b22; border:1px solid #30363d; border-radius:8px; padding:16px 18px; margin-bottom:12px; }
.kpi-card             { background:rgba(22, 27, 34, 0.5); border:1px solid rgba(255,255,255,0.05); border-radius:12px; padding:20px; position:relative; overflow:hidden; height:110px; box-shadow:0 4px 20px rgba(0,0,0,0.15); transition:all 0.3s ease; }
.kpi-card:hover       { transform:translateY(-2px); border-color:rgba(88,166,255,0.25); }
.sum-num              { font-size:28px; font-weight:700; }
.sum-lbl              { font-size:11px; color:#8b949e; text-transform:uppercase; letter-spacing:.5px; margin-top:2px; }

/* rec cards */
.rc                   { border-radius:8px; padding:11px 14px; margin-bottom:8px; font-size:13px; line-height:1.5; }
.rc-warn              { background:#1a0e0e; border-left:3px solid #f85149; }
.rc-caut              { background:#1a1500; border-left:3px solid #d29922; }
.rc-act               { background:#0d1f33; border-left:3px solid #58a6ff; }
.rc-info              { background:#13111f; border-left:3px solid #a78bfa; }
.rc-ok                { background:#0a1f12; border-left:3px solid #3fb950; }

/* auto-detect alert */
.auto-blk             { background:#1a0b0b; border:1px solid #f8514940; border-radius:8px; padding:12px 14px; margin-bottom:8px; }
.auto-risk            { background:#1a1200; border:1px solid #d2992240; border-radius:8px; padding:12px 14px; margin-bottom:8px; }

/* badges */
.badge                { display:inline-block; padding:2px 8px; border-radius:12px; font-size:11px; font-weight:600; }
.p-crit               { background:#3d1a1a; color:#ff6b6b; border:1px solid #ff6b6b40; }
.p-high               { background:#3d2e0a; color:#fbbf24; border:1px solid #fbbf2440; }
.p-med                { background:#0d2239; color:#60a5fa; border:1px solid #60a5fa40; }
.p-low                { background:#1a2039; color:#818cf8; border:1px solid #818cf840; }
.s-done               { background:#0a2a18; color:#3fb950; border:1px solid #3fb95040; }
.s-prog               { background:#0d2239; color:#58a6ff; border:1px solid #58a6ff40; }
.s-todo               { background:#21262d; color:#8b949e; border:1px solid #30363d; }
.s-blk                { background:#3d1a1a; color:#ff6b6b; border:1px solid #ff6b6b40; }
.s-auto               { background:#3d1a3d; color:#e879f9; border:1px solid #e879f940; }

/* activity */
.act-row              { display:flex; gap:10px; padding:9px 0; border-bottom:1px solid #21262d; align-items:flex-start; }

/* auth page */
.auth-brand            { margin-bottom:18px; }
.auth-title-center    { color:#e6edf3; font-size:30px; font-weight:700; text-align:center; margin-bottom:10px; letter-spacing:-.4px; }
.auth-sub-center      { color:#8b949e; font-size:13px; text-align:center; margin-bottom:0; }
</style>
""", unsafe_allow_html=True)

# ── DB & ML init ──────────────────────────────────────────────────────────────
# ADD THIS BELOW EXISTING CODE
@st.cache_resource(show_spinner="Initializing database connection...")
def init_db():
    conn = get_conn()
    create_tables(conn)
    seed_data(conn)
    conn.close()
    return True

init_db()

# ADD THIS BELOW EXISTING CODE
if not st.session_state.get("user"):
    if cookies.get("session_token"):
        delete_cookie("session_token")
    render_auth_gate()
    st.stop()

# Cookie synchronization for authenticated session
cookies = get_cookies()
if st.session_state.get("user"):
    # Write session cookie if missing
    if not cookies.get("session_token"):
        token = create_session_token(
            st.session_state["user"],
            st.session_state["role"],
            st.session_state.get("user_name", "")
        )
        set_cookie("session_token", token)
    
    # Write/update tab cookie if changed
    if cookies.get("session_tab") != st.session_state.tab:
        set_cookie("session_tab", st.session_state.tab)

@st.cache_resource(show_spinner=False)
def get_base_models():
    return RiskDetector(), AutoBlockerDetector(), VelocityForecaster(), RecommendationEngine()

def get_predictor_for_project(project_id):
    """Train a fresh predictor only on this project's historical tasks."""
    p = CompletionTimePredictor()
    conn = get_conn()
    df = pd.read_sql_query("""
        SELECT t.story_points, t.estimated_hours, t.actual_hours,
               t.priority, tm.velocity_avg, s.planned_points, s.completed_points
        FROM tasks t
        JOIN sprints s ON t.sprint_id = s.id
        JOIN team_members tm ON t.assignee_id = tm.id
        WHERE t.actual_hours IS NOT NULL
          AND s.status = 'Completed'
          AND s.project_id = %s
    """, conn, params=[project_id])
    conn.close()
    count = len(df)
    if count >= 5:
        p.train_from_df(df)
    return p, count

risk_det, auto_det, vel_fc, rec_eng = get_base_models()

# ── Matplotlib Chart Caching Helpers ──────────────────────────────────────────
@st.cache_data(ttl=300, show_spinner=False)
def get_status_donut_chart(done, inprog, todo, blk):
    fig, ax = plt.subplots(figsize=(3, 3), facecolor='none')
    szs = [done, inprog, todo, blk]
    cls = ['#3fb950', '#58a6ff', '#8b949e', '#f85149']
    lbls = ['Done', 'In Progress', 'To Do', 'Blocked']
    nz = [(s, c, l) for s, c, l in zip(szs, cls, lbls) if s > 0]
    if nz:
        s2, c2, _ = zip(*nz)
        ax.pie(s2, colors=c2, startangle=90, wedgeprops=dict(width=0.45, edgecolor='#0d1117', linewidth=2))
        total = sum(szs)
        ax.text(0, 0, str(total), ha='center', va='center', fontsize=20, fontweight='bold', color='#e6edf3')
    ax.set_facecolor('none')
    fig.patch.set_alpha(0)
    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', transparent=True, dpi=120)
    plt.close(fig)
    return buf.getvalue()

@st.cache_data(ttl=300, show_spinner=False)
def get_counts_bar_chart(todo, inprog, done):
    fig_counts, ax_counts = plt.subplots(figsize=(3.2, 1.4), facecolor="none")
    labels = ["Pending", "In Progress", "Completed"]
    values = [todo, inprog, done]
    ax_counts.bar(labels, values, color=["#8b949e", "#58a6ff", "#3fb950"])
    ax_counts.set_facecolor("#0d1117")
    fig_counts.patch.set_alpha(0)
    ax_counts.tick_params(colors="#8b949e", labelsize=9)
    ax_counts.spines[:].set_color("#21262d")
    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', transparent=True, dpi=120)
    plt.close(fig_counts)
    return buf.getvalue()

@st.cache_data(ttl=300, show_spinner=False)
def get_importance_chart(imps_tuple):
    fig4, ax4 = plt.subplots(figsize=(6, 2.5), facecolor='none')
    features = [x[0] for x in imps_tuple]
    importances = [x[1] for x in imps_tuple]
    ax4.barh(features, importances, color='#58a6ff', height=0.5)
    ax4.set_facecolor('#0d1117')
    fig4.patch.set_alpha(0)
    ax4.tick_params(colors='#8b949e', labelsize=9)
    ax4.spines[:].set_color('#21262d')
    ax4.set_xlabel("Importance", color='#8b949e', fontsize=9)
    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', transparent=True, dpi=120)
    plt.close(fig4)
    return buf.getvalue()

@st.cache_data(ttl=300, show_spinner=False)
def get_velocity_chart(planned_points_tuple, actual_points_tuple, sprint_names_tuple, forecast):
    fig5, ax5 = plt.subplots(figsize=(8, 3), facecolor='none')
    x = range(len(actual_points_tuple))
    ax5.bar(x, planned_points_tuple, color='#21262d', label='Planned', width=0.4, align='edge')
    ax5.bar([i - 0.4 for i in x], actual_points_tuple, color='#58a6ff', label='Actual', width=0.4, align='edge')
    ax5.axhline(forecast, color='#3fb950', linestyle='--', linewidth=1.5, label=f'Forecast {forecast}')
    ax5.set_xticks(x)
    ax5.set_xticklabels(sprint_names_tuple, color='#8b949e', fontsize=9)
    ax5.set_facecolor('#0d1117')
    fig5.patch.set_alpha(0)
    ax5.tick_params(colors='#8b949e')
    ax5.spines[:].set_color('#21262d')
    ax5.legend(facecolor='#161b22', edgecolor='#30363d', labelcolor='#8b949e', fontsize=9)
    buf = io.BytesIO()
    plt.savefig(buf, format='png', bbox_inches='tight', transparent=True, dpi=120)
    plt.close(fig5)
    return buf.getvalue()

@st.cache_data(ttl=60, show_spinner=False)
def get_sprint_scan_cached(sprint_id):
    tdf_s = load_tasks(sprint_id=sprint_id)
    if tdf_s.empty:
        return {"tasks": [], "remaining_days": 0, "auto_blocked_count": 0, "delayed_count": 0, "at_risk_count": 0, "total_delay_days": 0}
    df_sprint = qry("SELECT * FROM sprints WHERE id = %s", [sprint_id])
    if df_sprint.empty:
        return {"tasks": [], "remaining_days": 0, "auto_blocked_count": 0, "delayed_count": 0, "at_risk_count": 0, "total_delay_days": 0}
    sprint_dict = df_sprint.iloc[0].to_dict()
    return auto_det.scan_sprint(tdf_s.to_dict("records"), sprint_dict)

# ── DB helpers ────────────────────────────────────────────────────────────────
def qry(sql, params=()):
    if isinstance(params, list):
        params = tuple(params)
    elif not isinstance(params, tuple):
        params = (params,)
    return _qry_cached(sql, params)

@st.cache_data(ttl=60, show_spinner=False)
def _qry_cached(sql, params):
    c = get_conn()
    df = pd.read_sql_query(sql, c, params=list(params))
    c.close()
    return df

def exe(sql, params=()):
    c = get_conn()
    cur = c.cursor()
    sql_strip = sql.strip().upper()
    is_insert = sql_strip.startswith("INSERT")
    
    if is_insert and "RETURNING" not in sql_strip:
        query = sql + " RETURNING id"
        cur.execute(query, list(params))
        c.commit()
        lid = cur.fetchone()[0]
    else:
        cur.execute(sql, list(params))
        c.commit()
        try:
            lid = cur.fetchone()[0] if cur.description else None
        except Exception:
            lid = None
            
    cur.close()
    c.close()
    
    # Invalidate Streamlit cache on write
    st.cache_data.clear()
    
    # Reconcile sprints if tasks or sprints table is changed
    if "TASKS" in sql_strip or "SPRINTS" in sql_strip:
        try:
            reconcile_sprint_points()
        except Exception as e:
            print(f"Reconciliation failed: {e}")
            
    return lid

def load_projects():
    role = st.session_state.get("role")
    user = st.session_state.get("user")
    if not user:
        return pd.DataFrame()
    if role == "Manager":
        # Managers see their own projects + legacy projects with no creator
        return qry("SELECT * FROM projects WHERE created_by = %s OR created_by IS NULL ORDER BY id", [user])
    else:
        # Members see projects where their email matches
        df_email = qry("""
            SELECT DISTINCT p.*
            FROM projects p
            JOIN team_members tm ON p.id = tm.project_id
            WHERE LOWER(tm.email) = LOWER(%s)
            ORDER BY p.id
        """, [user])
        
        if not df_email.empty:
            return df_email
            
        # Fallback for legacy database rows where tm.email is NULL:
        # Load all projects and check using name-based heuristic
        df_proj = qry("SELECT * FROM projects ORDER BY id")
        if df_proj.empty:
            return df_proj
            
        df_team = qry("SELECT project_id, name, email FROM team_members")
        local = str(user).split("@")[0].lower().replace(".", "").replace("_", "")
        
        matched_project_ids = []
        for _, row in df_team.iterrows():
            if pd.notna(row.get("email")) and str(row["email"]).strip() != "":
                if str(row["email"]).lower() == str(user).lower():
                    matched_project_ids.append(row["project_id"])
            else:
                name_clean = str(row["name"]).lower().replace(".", "").replace("_", "")
                import re
                name_clean = re.sub(r"[^a-z0-9]", "", name_clean)
                if local in name_clean:
                    matched_project_ids.append(row["project_id"])
                    
        return df_proj[df_proj["id"].isin(matched_project_ids)]

def load_sprints(pid=None, status=None):
    s="SELECT * FROM sprints WHERE 1=1"; p=[]
    if pid:    s+=" AND project_id=%s"; p.append(pid)
    if status: s+=" AND status=%s";     p.append(status)
    return qry(s+" ORDER BY id", p)
def load_active(pid):
    df=qry("SELECT * FROM sprints WHERE status='Active' AND project_id=%s ORDER BY id DESC LIMIT 1",[pid])
    return df.iloc[0].to_dict() if not df.empty else None

def apply_role_task_filter(df):
    if df is None or df.empty:
        return df
    role = st.session_state.get("role")
    user = st.session_state.get("user")
    user_name = st.session_state.get("user_name")
    if role == "Member" and user:
        # Try strict email match first
        if "assignee_email" in df.columns:
            email_matches = df[df["assignee_email"].fillna("").str.lower() == str(user).lower()]
            if not email_matches.empty:
                return email_matches
                
        # Fallback to name heuristic for legacy tasks where email is NULL
        if "assignee_name" in df.columns:
            if user_name:
                exact = df[df["assignee_name"] == user_name]
                if not exact.empty:
                    return exact
            local = str(user).split("@")[0].lower().replace(".", "").replace("_", "")
            import re
            return df[
                df["assignee_name"].fillna("").str.lower().str.replace(r"[^a-z0-9]", "", regex=True).str.contains(local)
            ]
        if "assigned_to" in df.columns:
            return df[df["assigned_to"] == user]
    return df

def load_tasks(sprint_id=None, project_id=None, include_planning=False):
    s="""SELECT t.*,tm.name as assignee_name,tm.email as assignee_email,tm.velocity_avg,tm.avatar_color,sp.status as sprint_status
         FROM tasks t
         LEFT JOIN team_members tm ON t.assignee_id=tm.id
         LEFT JOIN sprints sp ON t.sprint_id=sp.id
         WHERE 1=1"""
    p=[]
    if sprint_id:  s+=" AND t.sprint_id=%s";  p.append(sprint_id)
    if project_id: s+=" AND t.project_id=%s"; p.append(project_id)
    # Hide tasks from not-started sprints unless explicitly requested.
    if not include_planning:
        s+=" AND (t.sprint_id IS NULL OR sp.status IS NULL OR sp.status!='Planning')"
    tdf = qry(s,p)
    return apply_role_task_filter(tdf)
def load_team(pid):     return qry("SELECT * FROM team_members WHERE project_id=%s ORDER BY id",[pid])
def load_metrics(pid):  return qry("""SELECT sm.*,s.name as sprint_name,s.planned_points
    FROM sprint_metrics sm JOIN sprints s ON sm.sprint_id=s.id
    WHERE s.project_id=%s ORDER BY sm.sprint_id""",[pid])
def load_activity(pid, n=20): return qry("SELECT * FROM activity_log WHERE project_id=%s ORDER BY id DESC LIMIT %s",[pid,n])

def reconcile_sprint_points(pid=None):
    c = get_conn()
    cur = c.cursor()

    if pid:
        # Auto-calculate completed points for all sprints
        cur.execute("""
            UPDATE sprints
            SET completed_points = COALESCE((
                SELECT SUM(t.story_points)
                FROM tasks t
                WHERE t.sprint_id = sprints.id
                AND t.status='Done'
            ), 0)
            WHERE project_id = %s
        """, [pid])

        # Auto-calculate planned points only for Active and Planning sprints
        # (Completed sprints preserve their historical planned_points)
        cur.execute("""
            UPDATE sprints
            SET planned_points = COALESCE((
                SELECT SUM(t.story_points)
                FROM tasks t
                WHERE t.sprint_id = sprints.id
            ), 0)
            WHERE project_id = %s AND status != 'Completed'
        """, [pid])
    else:
        # Auto-calculate completed points for all sprints across all projects
        cur.execute("""
            UPDATE sprints
            SET completed_points = COALESCE((
                SELECT SUM(t.story_points)
                FROM tasks t
                WHERE t.sprint_id = sprints.id
                AND t.status='Done'
            ), 0)
        """)

        # Auto-calculate planned points only for Active and Planning sprints across all projects
        cur.execute("""
            UPDATE sprints
            SET planned_points = COALESCE((
                SELECT SUM(t.story_points)
                FROM tasks t
                WHERE t.sprint_id = sprints.id
            ), 0)
            WHERE status != 'Completed'
        """)

    c.commit()
    cur.close()
    c.close()

def member_opts(tdf):
    d={None:"— Unassigned —"}
    d.update({r["id"]:f"{r['name']} ({r['role']})" for _,r in tdf.iterrows()})
    return d

def pbadge(p):
    m={"Critical":"p-crit","High":"p-high","Medium":"p-med","Low":"p-low"}
    i={"Critical":"🔴","High":"🟡","Medium":"🔵","Low":"⚪"}
    return f'<span class="badge {m.get(p,"p-med")}">{i.get(p,"")} {p}</span>'

def sbadge(s, auto=False):
    if auto: return '<span class="badge s-auto">🤖 Auto-Flagged</span>'
    m={"Done":"s-done","In Progress":"s-prog","Todo":"s-todo","Blocked":"s-blk"}
    i={"Done":"●","In Progress":"◑","Todo":"○","Blocked":"✕"}
    return f'<span class="badge {m.get(s,"s-todo")}">{i.get(s,"")} {s}</span>'

def av_html(name, color="#3b82f6", size=26):
    try:
        # Handle NaN, None, empty strings
        if pd.isna(name) or not str(name).strip():
            initial = "?"
        else:
            initial = str(name).strip()[0].upper()

        # Safe fallback color
        if pd.isna(color) or not str(color).strip():
            color = "#8b949e"

        return (
            f'<div style="width:{size}px;height:{size}px;'
            f'border-radius:50%;background:{color};'
            f'display:inline-flex;align-items:center;'
            f'justify-content:center;'
            f'font-size:{int(size*.4)}px;'
            f'font-weight:700;color:#fff;'
            f'flex-shrink:0">{initial}</div>'
        )

    except Exception:
        return (
            f'<div style="width:{size}px;height:{size}px;'
            f'border-radius:50%;background:#8b949e;'
            f'display:inline-flex;align-items:center;'
            f'justify-content:center;'
            f'color:#fff;font-weight:700;'
            f'flex-shrink:0">?</div>'
        )

def to_local_dt(dt_val):
    try:
        if pd.isna(dt_val):
            return None

        dt = pd.to_datetime(dt_val)

        # Ensure timezone aware
        if dt.tzinfo is None:
            dt = dt.tz_localize("UTC")

        # Convert UTC → India time
        return dt.tz_convert(ZoneInfo("Asia/Kolkata"))

    except Exception:
        return None

def chart_buf(fig):
    buf=io.BytesIO(); plt.savefig(buf,format='png',bbox_inches='tight',transparent=True,dpi=120)
    buf.seek(0); plt.close(fig); return buf

# ── SIDEBAR — project only, no nav ───────────────────────────────────────────
with st.sidebar:
    st.markdown("""
    <div style="padding:12px 2px 14px">
      <div style="display:flex;align-items:center;gap:10px;margin-bottom:8px">
        <div style="width:40px;height:40px;border-radius:10px;flex-shrink:0;
             background:linear-gradient(135deg,#2563eb,#7c3aed);
             box-shadow:0 0 18px #2563eb55;
             display:flex;align-items:center;justify-content:center;font-size:22px">⚡</div>
        <div>
          <div style="color:#e6edf3;font-weight:700;font-size:17px;letter-spacing:-.3px">SprintAI</div>
          <div style="color:#8b949e;font-size:10px;letter-spacing:.3px"></div>
        </div>
      </div>
      <div style="background:linear-gradient(90deg,#2563eb18,#7c3aed18);
           border:1px solid #2563eb30;border-radius:6px;
           padding:5px 10px;font-size:11px;color:#8b949e;text-align:center;letter-spacing:.2px">
        🤖 AI-Powered Agile Manager
      </div>
    </div>""", unsafe_allow_html=True)
    st.divider()

    projects = load_projects()
    if not projects.empty:
        sel_p    = st.selectbox("📁 Project", projects["name"].tolist(), label_visibility="visible")
        prow     = projects[projects["name"]==sel_p].iloc[0]
        proj_id  = int(prow["id"]); proj_name = prow["name"]
    else:
        proj_id = None; proj_name = "—"

    # ADD THIS BELOW EXISTING CODE
    if st.session_state.get("role") == "Manager":
        with st.expander("＋ New Project"):
            with st.form("np", clear_on_submit=True):
                n=st.text_input("Name *"); d=st.text_area("Description",height=50)
                if st.form_submit_button("Create",type="primary"):
                    if n.strip():
                        user_email = st.session_state.get("user")
                        owner_df = qry("SELECT id FROM profiles WHERE LOWER(email) = LOWER(%s)", [user_email])
                        if not owner_df.empty:
                            owner_id = owner_df.iloc[0]["id"]
                            exe("INSERT INTO projects(name,description,created_by,owner_id)VALUES(%s,%s,%s,%s)",(n.strip(),d,user_email,owner_id))
                        else:
                            exe("INSERT INTO projects(name,description,created_by)VALUES(%s,%s,%s)",(n.strip(),d,user_email))
                        st.rerun()
                    else: st.error("Name required")
    else:
        st.caption("Member access: project creation is restricted.")

    if proj_id:
        st.divider()
        act_s = load_active(proj_id)
        if act_s:
            pp=act_s["planned_points"]; cp=act_s["completed_points"]; pct=cp/max(pp,1)
            st.markdown(f'<div style="color:#3fb950;font-size:12px;font-weight:600;margin-bottom:4px">🟢 {act_s["name"]}</div>', unsafe_allow_html=True)
            st.caption(act_s.get("goal","") or "No goal set")
            st.progress(min(pct,1.0), text=f"{cp}/{pp} pts · {pct*100:.0f}%")
            # Auto-detection summary in sidebar (Phase 2 Cached Scan) - only run on metrics-heavy tabs
            if st.session_state.get("tab") in ["Dashboard", "AI Insights", "Reports", "Projects", "Board"]:
                scan = get_sprint_scan_cached(int(act_s["id"]))
                if scan["auto_blocked_count"] or scan["delayed_count"] or scan["at_risk_count"]:
                    st.markdown(f"""
                    <div style="background:#1a0b0b;border:1px solid #f8514940;border-radius:6px;padding:8px 10px;margin-top:8px">
                      <div style="color:#f85149;font-size:11px;font-weight:700">🤖 AI DETECTED</div>
                      <div style="color:#c9d1d9;font-size:11px;margin-top:4px">
                        {f'⛔ {scan["auto_blocked_count"]} auto-flagged' if scan["auto_blocked_count"] else ''}
                        {f'<br>⏰ {scan["delayed_count"]} delayed' if scan["delayed_count"] else ''}
                        {f'<br>⚠️ {scan["at_risk_count"]} at risk' if scan["at_risk_count"] else ''}
                      </div>
                    </div>""", unsafe_allow_html=True)
        else:
            st.caption("⚪ No active sprint")

    st.divider()
    st.caption("Python · Streamlit · Scikit-learn · SQLite")
    # ADD THIS BELOW EXISTING CODE
    st.caption(f"👤 {st.session_state.get('user')} ({st.session_state.get('role')})")
    if st.button("🚪 Logout", use_container_width=True):
        st.session_state["logout_triggered"] = True
        logout_user()
        delete_cookie("session_token")
        delete_cookie("session_tab")
        st.rerun()

# ── Guard ─────────────────────────────────────────────────────────────────────
if not proj_id:
    st.markdown("""
    <div style="display:flex;flex-direction:column;align-items:center;justify-content:center;height:70vh;gap:12px">
      <div style="font-size:48px">⚡</div>
      <div style="color:#e6edf3;font-size:22px;font-weight:700">Welcome to SprintAI</div>
      <div style="color:#8b949e">Create your first project in the sidebar.</div>
    </div>""", unsafe_allow_html=True)
    st.stop()

# Reconciled in the background on data modifications, no need to run on every page load.

# ══════════════════════════════════════════════════════════════════════════════
#  TOP NAV — project header + clickable tabs
# ══════════════════════════════════════════════════════════════════════════════
TABS = ["Dashboard","Projects","Board","Sprints","Task Creation","Team","AI Insights","Reports","Daily Updates"]
TAB_ICONS = {"Dashboard":"📊","Projects":"📋","Board":"🗂️","Sprints":"🏃",
             "Task Creation":"🗒️","Team":"👥","AI Insights":"🔮","Reports":"📄","Daily Updates":"📝"}

# ── Project name above tabs (like Jira) ──────────────────────────────────────
st.markdown(f"""
<div style="background:#161b22;padding:8px 24px 0 24px;border-bottom:none">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:14px">
    <div style="width:28px;height:28px;border-radius:6px;
         background:linear-gradient(135deg,#f97316,#ef4444);
         display:flex;align-items:center;justify-content:center;font-size:14px;flex-shrink:0">🚀</div>
    <span style="color:#e6edf3;font-size:20px;font-weight:700">{proj_name}</span>
  </div>
</div>""", unsafe_allow_html=True)

# Tab bar — using columns + buttons
tab_cols = st.columns(len(TABS))
for i, t in enumerate(TABS):
    with tab_cols[i]:
        is_active = (st.session_state.tab == t)
        label = f"{TAB_ICONS[t]} {t}"
        if is_active:
            st.markdown(f"""
            <div style="text-align:center;padding:10px 4px 8px;
                 border-bottom:2px solid #58a6ff;background:#161b22;
                 color:#58a6ff;font-size:12px;font-weight:600;
                 cursor:pointer;white-space:nowrap">{label}</div>""",
                unsafe_allow_html=True)
        else:
            if st.button(label, key=f"tab_{t}", use_container_width=True):
                if st.session_state.get("tab") != t:
                    st.session_state.tab = t
                    st.rerun()

st.markdown('<div style="background:#30363d;height:1px;margin-bottom:0"></div>', unsafe_allow_html=True)

page = st.session_state.tab

# ══════════════════════════════════════════════════════════════════════════════
#  📊 SUMMARY
# ══════════════════════════════════════════════════════════════════════════════
if page == "Dashboard":
    st.markdown('<div class="page-wrap">', unsafe_allow_html=True)

    act   = load_active(proj_id)
    tdf   = load_tasks(sprint_id=act["id"]) if act else load_tasks(project_id=proj_id)
    tasks = tdf.to_dict("records")
    team  = load_team(proj_id)
    acts  = load_activity(proj_id, 20)
    mets  = load_metrics(proj_id)

    total  = len(tasks)
    done   = sum(1 for t in tasks if t["status"]=="Done")
    inprog = sum(1 for t in tasks if t["status"]=="In Progress")
    todo   = sum(1 for t in tasks if t["status"]=="Todo")
    blk    = sum(1 for t in tasks if t["status"]=="Blocked")

    # Auto-detection scan (Phase 2 Cached Scan)
    if act:
        scan = get_sprint_scan_cached(int(act["id"]))
        ab   = scan["auto_blocked_count"]
        dl   = scan["delayed_count"]
        ar   = scan["at_risk_count"]
        if ab or dl or ar:
            st.markdown(f"""
            <div style="background:#1a0b0b;border:1px solid #f8514960;border-radius:8px;
                 padding:12px 16px;margin-bottom:16px;display:flex;gap:24px;align-items:center">
              <span style="color:#f85149;font-weight:700;font-size:14px">🤖 AI Auto-Detection</span>
              {'<span style="color:#f85149;font-size:13px">⛔ '+str(ab)+' task(s) auto-flagged as blocked</span>' if ab else ''}
              {'<span style="color:#d29922;font-size:13px">⏰ '+str(dl)+' task(s) delayed</span>' if dl else ''}
              {'<span style="color:#fbbf24;font-size:13px">⚠️ '+str(ar)+' at risk</span>' if ar else ''}
            </div>""", unsafe_allow_html=True)

    # KPI row
    if act:
        hlth = compute_sprint_health(tasks, act)
        health_grade = hlth["grade"]
        health_score = hlth["score"]
        health_color = {"A":"#3fb950","B":"#d29922","C":"#f0883e","D":"#f85149","F":"#f85149"}.get(health_grade, "#8b949e")
        active_sprint_name = act["name"]
    else:
        health_grade = "—"
        health_score = "0"
        health_color = "#8b949e"
        active_sprint_name = "No Active Sprint"

    k1, k2, k3, k4 = st.columns(4)
    k1.markdown(f"""
    <div class="kpi-card">
        <div style="font-size: 11px; color: #8b949e; text-transform: uppercase; letter-spacing: 0.5px; font-weight: 600;">Total Tasks</div>
        <div style="font-family: 'Outfit', sans-serif; font-size: 32px; font-weight: 700; color: #ffffff; margin-top: 8px;">{total}</div>
        <div style="position: absolute; right: 15px; bottom: 15px; font-size: 26px; opacity: 0.25;">📋</div>
    </div>
    """, unsafe_allow_html=True)
    k2.markdown(f"""
    <div class="kpi-card">
        <div style="font-size: 11px; color: #8b949e; text-transform: uppercase; letter-spacing: 0.5px; font-weight: 600;">Active Sprint</div>
        <div style="font-family: 'Outfit', sans-serif; font-size: 22px; font-weight: 700; color: #3fb950; margin-top: 16px; overflow: hidden; text-overflow: ellipsis; white-space: nowrap;">{active_sprint_name}</div>
        <div style="position: absolute; right: 15px; bottom: 15px; font-size: 26px; opacity: 0.25;">🏃</div>
    </div>
    """, unsafe_allow_html=True)
    k3.markdown(f"""
    <div class="kpi-card">
        <div style="font-size: 11px; color: #8b949e; text-transform: uppercase; letter-spacing: 0.5px; font-weight: 600;">Team Size</div>
        <div style="font-family: 'Outfit', sans-serif; font-size: 32px; font-weight: 700; color: #58a6ff; margin-top: 8px;">{len(team)} <span style="font-size:14px; font-weight:400; color:#8b949e;">members</span></div>
        <div style="position: absolute; right: 15px; bottom: 15px; font-size: 26px; opacity: 0.25;">👥</div>
    </div>
    """, unsafe_allow_html=True)
    k4.markdown(f"""
    <div class="kpi-card">
        <div style="font-size: 11px; color: #8b949e; text-transform: uppercase; letter-spacing: 0.5px; font-weight: 600;">Sprint Health</div>
        <div style="font-family: 'Outfit', sans-serif; font-size: 32px; font-weight: 700; color: {health_color}; margin-top: 8px;">{health_grade} <span style="font-size: 14px; font-weight: 400; color: #8b949e;">({health_score}/100)</span></div>
        <div style="position: absolute; right: 15px; bottom: 15px; font-size: 26px; opacity: 0.25;">🔮</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("<br>", unsafe_allow_html=True)

    left, right = st.columns([1.4,1])

    with left:
        # Status donut
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("**📊 Status Overview**")
        if total > 0:
            szs = [done, inprog, todo, blk]
            cls = ['#3fb950', '#58a6ff', '#8b949e', '#f85149']
            lbls = ['Done', 'In Progress', 'To Do', 'Blocked']
            cl,cr=st.columns([1,1.1])
            with cl:
                st.image(get_status_donut_chart(done, inprog, todo, blk), use_container_width=True)
            with cr:
                for s,c,l in zip(szs,cls,lbls):
                    if s>0:
                        st.markdown(f"""
                        <div style="display:flex;align-items:center;gap:8px;margin-bottom:9px">
                          <div style="width:10px;height:10px;border-radius:2px;background:{c}"></div>
                          <span style="color:#c9d1d9;font-size:13px;flex:1">{l}</span>
                          <span style="color:#8b949e;font-size:12px">{s} ({s/total*100:.0f}%)</span>
                        </div>""", unsafe_allow_html=True)
        else:
            st.caption("No tasks yet.")
        st.markdown('</div>', unsafe_allow_html=True)

        # Priority breakdown
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("**🎯 Priority Breakdown**")
        for pn,pc in [("Critical","#f85149"),("High","#d29922"),("Medium","#58a6ff"),("Low","#8b949e")]:
            cnt=sum(1 for t in tasks if t["priority"]==pn)
            if cnt:
                pct2=cnt/max(total,1)*100
                st.markdown(f"""
                <div style="margin-bottom:9px">
                  <div style="display:flex;justify-content:space-between;margin-bottom:3px">
                    <span style="color:#c9d1d9;font-size:12px">{pn}</span>
                    <span style="color:#8b949e;font-size:12px">{cnt} ({pct2:.0f}%)</span>
                  </div>
                  <div style="background:#21262d;border-radius:3px;height:5px">
                    <div style="width:{pct2}%;background:{pc};height:5px;border-radius:3px"></div>
                  </div>
                </div>""", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    with right:
        # AI sprint intelligence (Slimmed)
        if act:
            hlth  = compute_sprint_health(tasks,act)
            s_risk= risk_det.assess_sprint_risk(act,tasks)
            gc    = {"A":"#3fb950","B":"#d29922","C":"#f0883e","D":"#f85149","F":"#f85149"}.get(hlth["grade"],"#8b949e")
            rc    = {"low":"#3fb950","medium":"#d29922","high":"#f85149"}[s_risk["level"]]

            pp = act["planned_points"]
            cp = act["completed_points"]
            pct_cp = cp / max(pp, 1)

            st.markdown(f"""
            <div class="card" style="margin-bottom:10px">
              <div style="display:flex;justify-content:space-between;align-items:center;margin-bottom:12px">
                <span style="color:#e6edf3;font-weight:600;font-size:14px">🤖 AI Sprint Intelligence</span>
                <span style="background:{gc}20;color:{gc};border:1px solid {gc}40;border-radius:12px;padding:2px 10px;font-size:11px;font-weight:700">
                  Grade {hlth['grade']} · {hlth['score']}/100
                </span>
              </div>
              <div style="display:grid;grid-template-columns:1fr 1fr 1fr;gap:8px;margin-bottom:10px">
                <div style="background:#21262d;border-radius:6px;padding:10px;text-align:center">
                  <div style="color:{rc};font-size:13px;font-weight:700">{s_risk['level'].upper()}</div>
                  <div style="color:#8b949e;font-size:10px">Sprint Risk</div>
                </div>
                <div style="background:#21262d;border-radius:6px;padding:10px;text-align:center">
                  <div style="color:#58a6ff;font-size:13px;font-weight:700">{pct_cp*100:.0f}%</div>
                  <div style="color:#8b949e;font-size:10px">Completed Pts</div>
                </div>
                <div style="background:#21262d;border-radius:6px;padding:10px;text-align:center">
                  <div style="color:#{'f85149' if blk>0 else '3fb950'};font-size:13px;font-weight:700">{blk}</div>
                  <div style="color:#8b949e;font-size:10px">Blockers</div>
                </div>
              </div>
            </div>""", unsafe_allow_html=True)

            st.markdown("**🔔 Recommendation**")
            if blk > 0:
                st.markdown('<div class="rc rc-warn"><strong>⚠️ Too many blocked tasks detected</strong><br><span style="color:#8b949e">Address active blockers immediately to unblock team members.</span></div>', unsafe_allow_html=True)
            elif s_risk["level"] == "high":
                st.markdown('<div class="rc rc-act"><strong>💡 Reassign backend tasks</strong><br><span style="color:#8b949e">Workload is highly unbalanced. Reassign tasks to high-velocity members.</span></div>', unsafe_allow_html=True)
            else:
                st.markdown('<div class="rc rc-ok"><strong>✅ Sprint on track</strong><br><span style="color:#8b949e">No blockers or high risk detected. Keep up the good work!</span></div>', unsafe_allow_html=True)

        # Activity feed
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("**⚡ Recent Activity**")
        cmap={"K. Amulya":"#ec4899","M. Bhavani":"#3b82f6","B.G.L. Santhoshi":"#10b981","Mr. Sudheer":"#f59e0b"}
        if acts.empty:
            st.caption("No activity yet.")
        else:
            for _,a in acts.head(8).iterrows():
                actor=str(a.get("actor","System")); col=cmap.get(actor,"#8b949e")
                fld=a.get("field_changed",""); nv=a.get("new_value",""); tt=a.get("task_title",""); ts_dt=to_local_dt(a.get("created_at")); ts=ts_dt.strftime("%Y-%m-%d %H:%M") if ts_dt else str(a.get("created_at",""))[:16]
                desc=f'<b style="color:#58a6ff">{a.get("action","updated")}</b>'
                if fld and nv: desc+=f' <span style="color:#8b949e">{fld}</span> → <span style="color:#3fb950">{nv}</span>'
                if tt: desc+=f' on <b style="color:#c9d1d9">{tt}</b>'
                st.markdown(f"""
                <div class="act-row">
                  {av_html(actor,col,24)}
                  <div>
                    <span style="color:#c9d1d9;font-size:12px;font-weight:600">{actor}</span>
                    <span style="color:#8b949e;font-size:12px"> {desc}</span>
                    <div style="color:#8b949e;font-size:10px;margin-top:1px">{ts}</div>
                  </div>
                </div>""", unsafe_allow_html=True)
        st.markdown('</div>', unsafe_allow_html=True)

    # Velocity + Types row
    vc,tc=st.columns(2)
    with vc:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("**📈 Velocity Trend**")
        if not mets.empty:
            mv=mets["velocity"].max()
            for _,r in mets.iterrows():
                v=r["velocity"]; p=r["planned_points"]; c2="#3fb950" if v>=p else "#d29922"
                st.markdown(f"""
                <div style="margin-bottom:8px">
                  <div style="display:flex;justify-content:space-between;margin-bottom:2px">
                    <span style="color:#8b949e;font-size:11px">{r['sprint_name']}</span>
                    <span style="color:{c2};font-size:11px;font-weight:600">{v:.0f}/{p:.0f}pt</span>
                  </div>
                  <div style="background:#21262d;border-radius:3px;height:5px">
                    <div style="width:{min(v/max(mv,1)*100,100):.0f}%;background:{c2};height:5px;border-radius:3px"></div>
                  </div>
                </div>""", unsafe_allow_html=True)
        else:
            st.caption("Complete sprints to see trend.")
        st.markdown('</div>', unsafe_allow_html=True)

    with tc:
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("**📦 Types of Work**")
        itcol={"Story":"#58a6ff","Task":"#3fb950","Bug":"#f85149","Subtask":"#d29922","Epic":"#a78bfa"}
        if not tdf.empty and "issue_type" in tdf.columns:
            for it,cnt in tdf["issue_type"].value_counts().items():
                c2=itcol.get(it,"#8b949e")
                st.markdown(f"""
                <div style="display:flex;align-items:center;gap:8px;margin-bottom:8px">
                  <div style="width:8px;height:8px;border-radius:50%;background:{c2}"></div>
                  <span style="color:#c9d1d9;font-size:13px;flex:1">{it}</span>
                  <span style="color:#8b949e;font-size:12px">{cnt} ({cnt/max(total,1)*100:.0f}%)</span>
                </div>""", unsafe_allow_html=True)
        else:
            st.caption("No tasks yet.")
        st.markdown('</div>', unsafe_allow_html=True)

    st.markdown('</div>', unsafe_allow_html=True)  # page-wrap


# ══════════════════════════════════════════════════════════════════════════════
#  📋 LIST
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Projects":
    st.markdown('<div class="page-wrap">', unsafe_allow_html=True)
    sdf=load_sprints(pid=proj_id); team=load_team(proj_id)
    act=load_active(proj_id)

    fc1,fc2,fc3,fc4,fc5=st.columns(5)
    f_spr=fc1.selectbox("Sprint",  ["All"]+(sdf["name"].tolist() if not sdf.empty else []))
    f_st =fc2.selectbox("Status",  ["All","Todo","In Progress","Done","Blocked"])
    f_pr =fc3.selectbox("Priority",["All","Critical","High","Medium","Low"])
    f_as =fc4.selectbox("Assignee",["All"]+(team["name"].tolist() if not team.empty else []))
    f_risk=fc5.selectbox("Risk",   ["All","Auto-Flagged","Delayed","At Risk"])

    all_t=load_tasks(project_id=proj_id)

    # Apply filters
    if not all_t.empty:
        if f_spr!="All" and not sdf.empty:
            sid2=int(sdf[sdf["name"]==f_spr]["id"].values[0]); all_t=all_t[all_t["sprint_id"]==sid2]
        if f_st !="All": all_t=all_t[all_t["status"]==f_st]
        if f_pr !="All": all_t=all_t[all_t["priority"]==f_pr]
        if f_as !="All": all_t=all_t[all_t["assignee_name"]==f_as]

    # Run auto-detection on filtered tasks
    scan_data = {}
    if act and not all_t.empty:
        scanned = get_sprint_scan_cached(int(act["id"]))
        for t in scanned["tasks"]:
            scan_data[t["id"]] = t["auto_detection"]

    # Filter by risk
    if f_risk != "All" and scan_data:
        risk_filter_ids = set()
        for tid, det in scan_data.items():
            if f_risk == "Auto-Flagged" and det["auto_blocked"]:  risk_filter_ids.add(tid)
            if f_risk == "Delayed"      and det["delay_risk"]=="delayed": risk_filter_ids.add(tid)
            if f_risk == "At Risk"      and det["delay_risk"]=="at_risk": risk_filter_ids.add(tid)
        if risk_filter_ids:
            all_t = all_t[all_t["id"].isin(risk_filter_ids)]
        else:
            all_t = pd.DataFrame()

    st.caption(f"{len(all_t) if not all_t.empty else 0} items")

    if all_t.empty:
        st.info("No tasks match filters.")
    else:
        # Header row
        st.markdown("""
        <div style="display:grid;grid-template-columns:2fr 1fr 1fr 1fr 1fr 1fr 1fr;
             gap:6px;padding:8px 12px;background:#161b22;border:1px solid #30363d;
             border-radius:6px 6px 0 0;font-size:11px;color:#8b949e;font-weight:600;
             text-transform:uppercase;letter-spacing:.5px">
          <div>Title</div><div>Assignee</div><div>Priority</div><div>Status</div>
          <div>AI Risk</div><div>Points</div><div>Est. Hrs</div>
        </div>""", unsafe_allow_html=True)

        for _,t in all_t.iterrows():
            det  = scan_data.get(t["id"], {})
            ab   = det.get("auto_blocked", False)
            dr   = det.get("delay_risk","none")
            av   = t.get("assignee_name") or "Unassigned"
            avc  = t.get("avatar_color") or "#8b949e"
            reasons = det.get("reasons",[])

            risk_label = '🤖 <span style="color:#e879f9">AUTO-FLAGGED</span>' if ab else \
                         '⏰ <span style="color:#f85149">DELAYED</span>'     if dr=="delayed" else \
                         '⚠️ <span style="color:#d29922">AT RISK</span>'      if dr=="at_risk" else \
                         '🟢 <span style="color:#3fb950">OK</span>'
            border_col = "#f85149" if ab or dr=="delayed" else "#d29922" if dr=="at_risk" else "#21262d"

            blocker_html = f'<div style="color:#f85149;font-size:10px;margin-top:2px">⛔ {t["blocker_note"]}</div>' if t.get("blocker_note") else ''
            reason_html  = f'<div style="color:#d29922;font-size:10px;margin-top:2px">🤖 {reasons[0]}</div>' if reasons else ''

            st.markdown(
                f'<div style="display:grid;grid-template-columns:2fr 1fr 1fr 1fr 1fr 1fr 1fr;gap:6px;padding:12px;background:rgba(22, 27, 34, 0.45);border:1px solid rgba(255, 255, 255, 0.05);border-radius:6px;margin-bottom:8px;border-left:3px solid {border_col};font-size:13px">'
                f'<div style="color:#e6edf3;font-weight:500">{t["title"]}{blocker_html}{reason_html}</div>'
                f'<div style="display:flex;align-items:center;gap:5px">{av_html(av,avc,18)} <span style="color:#8b949e;font-size:11px">{av}</span></div>'
                f'<div>{pbadge(t["priority"])}</div>'
                f'<div>{sbadge(t["status"])}</div>'
                f'<div style="font-size:11px">{risk_label}</div>'
                f'<div style="color:#8b949e">{t.get("story_points",1)}pt</div>'
                f'<div style="color:#8b949e">{t.get("estimated_hours") or "—"}h</div>'
                f'</div>',
                unsafe_allow_html=True
            )

    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  🗂️ BOARD
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Board":
    st.markdown('<div class="page-wrap">', unsafe_allow_html=True)
    sdf=load_sprints(pid=proj_id); team=load_team(proj_id)
    if sdf.empty: st.warning("No sprints yet. Go to **Sprints** to create one."); st.stop()

    names=sdf["name"].tolist()
    def_idx=next((i for i,s in enumerate(sdf["status"]) if s=="Active"),0)
    sel=st.selectbox("Sprint",names,index=def_idx)
    srow=sdf[sdf["name"]==sel].iloc[0]; sid=int(srow["id"])
    if "board_sprint_id" not in st.session_state or st.session_state.board_sprint_id != sid:
        st.session_state.board_sprint_id = sid
        st.session_state.board_limits = {"Todo": 20, "In Progress": 20, "Done": 20, "Blocked": 20}
    st.caption(f"📅 {srow['start_date']} → {srow['end_date']}  ·  Status: **{srow['status']}**  ·  Goal: *{srow.get('goal') or '—'}*")

    tdf=load_tasks(sprint_id=sid); tasks=tdf.to_dict("records")

    # Auto-detect
    scan2={}
    if not tdf.empty:
        sc=get_sprint_scan_cached(sid)
        if sc["auto_blocked_count"] or sc["delayed_count"]:
            st.markdown(f"""
            <div style="background:#1a0b0b;border:1px solid #f8514960;border-radius:8px;
                 padding:10px 16px;margin-bottom:12px;display:flex;gap:20px;align-items:center">
              <span style="color:#f85149;font-weight:700">🤖 AI Detected:</span>
              {'<span style="color:#f85149;font-size:13px">⛔ '+str(sc["auto_blocked_count"])+' auto-flagged</span>' if sc["auto_blocked_count"] else ''}
              {'<span style="color:#d29922;font-size:13px">⏰ '+str(sc["delayed_count"])+' delayed</span>' if sc["delayed_count"] else ''}
              {'<span style="color:#fbbf24;font-size:13px">⚠️ '+str(sc["at_risk_count"])+' at risk</span>' if sc["at_risk_count"] else ''}
              <span style="color:#8b949e;font-size:12px">See AI Insights tab for details →</span>
            </div>""", unsafe_allow_html=True)
        for t in sc["tasks"]:
            scan2[t["id"]] = t["auto_detection"]

    with st.expander("✏️ Update Task Status / Log Hours"):
        if tdf.empty: st.caption("No tasks.")
        else:
            with st.form("bu"):
                titles=tdf["title"].tolist(); ids=tdf["id"].tolist()
                sel_t=st.selectbox("Task",titles); tid=ids[titles.index(sel_t)]
                cur=tdf[tdf["id"]==tid].iloc[0]
                uc1,uc2,uc3=st.columns(3)
                ns=uc1.selectbox("Status",["Todo","In Progress","Done","Blocked"],index=["Todo","In Progress","Done","Blocked"].index(cur["status"]))
                na=uc2.number_input("Actual Hours",0.0,200.0,float(cur.get("actual_hours") or 0.0),step=0.5)
                nb=uc3.text_input("Blocker Note",value=cur.get("blocker_note") or "")
                if st.form_submit_button("💾 Save",type="primary"):
                    exe("UPDATE tasks SET status=%s,actual_hours=%s,blocker_note=%s,updated_at=CURRENT_TIMESTAMP WHERE id=%s",
                        (ns,na or None,nb or None,tid))
                    if ns!=cur["status"]:
                        if ns=="Done":    exe("UPDATE sprints SET completed_points=completed_points+%s WHERE id=%s",(int(cur["story_points"]),sid))
                        elif cur["status"]=="Done": exe("UPDATE sprints SET completed_points=GREATEST(0,completed_points-%s) WHERE id=%s",(int(cur["story_points"]),sid))
                    log_activity(proj_id,"User","updated",sel_t,"status",cur["status"],int(cur["id"]),sid,sel_t)
                    st.success("Updated!"); st.rerun()

    st.markdown("<br>", unsafe_allow_html=True)
    COLS=["Todo","In Progress","Done","Blocked"]
    BORDS={"Todo":"#475569","In Progress":"#2563eb","Done":"#059669","Blocked":"#dc2626"}
    cws=st.columns(4)
    for idx,status in enumerate(COLS):
        col_t=[t for t in tasks if t["status"]==status]; pts=sum(t["story_points"] for t in col_t); bc=BORDS[status]
        with cws[idx]:
            st.markdown(f"""
            <div style="background:#161b22;border:1px solid {bc}50;border-top:3px solid {bc};
                 border-radius:8px;padding:8px 12px;margin-bottom:8px;
                 display:flex;justify-content:space-between;align-items:center">
              <span style="color:#e6edf3;font-weight:600;font-size:13px">{status}</span>
              <span style="background:{bc}25;border-radius:10px;padding:1px 8px;font-size:11px;color:#c9d1d9">{len(col_t)}·{pts}pt</span>
            </div>""", unsafe_allow_html=True)

            if not col_t:
                st.markdown('<div style="color:#30363d;text-align:center;padding:28px;font-size:12px;border:1px dashed #21262d;border-radius:6px">— empty —</div>',unsafe_allow_html=True)

            limit = st.session_state.board_limits.get(status, 20)
            displayed_tasks = col_t[:limit]

            for t in displayed_tasks:
                det2=scan2.get(t["id"],{})
                ab2=det2.get("auto_blocked",False); dr2=det2.get("delay_risk","none")
                ai_reasons=det2.get("reasons",[])
                bord_c = "#e879f9" if ab2 else "#f85149" if dr2 == "delayed" else "#d29922" if dr2 == "at_risk" else "#3b82f640"
                
                # AI Indicator Styling
                if ab2:
                    ai_banner='<div style="color:#e879f9;font-size:10.5px;margin-top:6px;font-weight:600;">🤖 AI: Auto-blocked</div>'
                elif dr2=="delayed":
                    ai_banner=f'<div style="color:#f85149;font-size:10.5px;margin-top:6px;font-weight:600;">⏰ Delayed ~{det2.get("delay_days",0)}d</div>'
                elif dr2=="at_risk":
                    ai_banner=f'<div style="color:#d29922;font-size:10.5px;margin-top:6px;font-weight:600;">⚠️ {ai_reasons[0] if ai_reasons else "At risk"}</div>'
                else:
                    ai_banner=""
                
                blk_note=f'<div style="color:#f85149;font-size:10.5px;margin-top:6px;padding-top:6px;border-top:1px solid rgba(255,255,255,0.05);font-weight:600;">⛔ {t["blocker_note"]}</div>' if t.get("blocker_note") else ""
                av=t.get("assignee_name") or "Unassigned"; avc=t.get("avatar_color") or "#8b949e"
                st.markdown(
                    f'<div style="background:#161b22;border:1px solid {bord_c};border-radius:6px;padding:10px 12px;margin-bottom:6px">'
                    f'<div style="color:#e6edf3;font-weight:500;font-size:13px;margin-bottom:5px">{t["title"]}</div>'
                    f'<div style="display:flex;gap:4px;flex-wrap:wrap;margin-bottom:5px">{pbadge(t["priority"])}<span class="badge s-todo" style="font-size:10px">{t["story_points"]}pt</span></div>'
                    f'<div style="display:flex;align-items:center;gap:5px">{av_html(av,avc,18)} <span style="color:#8b949e;font-size:11px">{av}</span></div>'
                    f'{ai_banner}{blk_note}'
                    f'</div>',
                    unsafe_allow_html=True
                )

            if len(col_t) > limit:
                if st.button(f"Show More (+{len(col_t) - limit})", key=f"more_{status}_{sid}", use_container_width=True):
                    st.session_state.board_limits[status] = limit + 20
                    st.rerun()
                
                # Priority Border Colors
                p_c = {"Critical": "#f85149", "High": "#d29922", "Medium": "#58a6ff", "Low": "#8b949e"}.get(t["priority"], "#8b949e")
                due_date_str = f' · 📅 {t["due_date"]}' if t.get("due_date") else ""
                
                st.markdown(f"""
                <div class="card" style="border-left: 4px solid {p_c}; margin-bottom: 8px; padding: 12px 14px; background: rgba(22, 27, 34, 0.45); border-color: rgba(255, 255, 255, 0.05);">
                    <div style="color:#e6edf3; font-weight:600; font-size:12.5px; line-height:1.4; margin-bottom:6px;">{t["title"]}</div>
                    <div style="display:flex; justify-content:space-between; align-items:center; margin-top:8px;">
                        <div style="display:flex; align-items:center; gap:6px;">
                            {av_html(av, avc, 18)}
                            <span style="color:#8b949e; font-size:11px;">{av}</span>
                        </div>
                        <span style="background:rgba(255,255,255,0.04); border:1px solid rgba(255,255,255,0.08); border-radius:4px; padding:2px 6px; font-size:10px; color:#c9d1d9; font-weight:600;">{t["story_points"]}pt{due_date_str}</span>
                    </div>
                    {ai_banner}
                    {blk_note}
                </div>
                """, unsafe_allow_html=True)
    st.markdown('</div>', unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  🏃 SPRINTS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Sprints":
    st.markdown('<div class="page-wrap">', unsafe_allow_html=True)
    if st.session_state.get("role") == "Manager":
        with st.expander("➕ Create New Sprint",expanded=True):
            with st.form("cs", clear_on_submit=True):
                st.markdown("#### New Sprint")
                sc1,sc2=st.columns(2)
                sn=sc1.text_input("Sprint Name *",placeholder="Sprint 1"); sg=sc1.text_input("Sprint Goal",placeholder="e.g. Complete auth module")
                ss=sc2.date_input("Start Date",value=date.today()); se=sc2.date_input("End Date",value=date.today()+timedelta(days=14))
                sp=sc2.number_input("Planned Story Points",0,300,40)
                if st.form_submit_button("🚀 Create Sprint",type="primary"):
                    if not sn.strip(): st.error("Name required.")
                    elif se<=ss: st.error("End must be after start.")
                    else:
                        ex=load_sprints(pid=proj_id)
                        if not ex.empty and sn.strip() in ex["name"].tolist(): st.error("Name exists.")
                        else:
                            exe("INSERT INTO sprints(project_id,name,goal,start_date,end_date,status,planned_points,completed_points)VALUES(%s,%s,%s,%s,%s,%s,%s,%s)",
                                (proj_id,sn.strip(),sg.strip(),str(ss),str(se),"Planning",sp,0))
                            log_activity(proj_id,"User","created sprint",sn.strip())
                            st.success(f"Sprint '{sn}' created!"); st.rerun()
    else:
        st.caption("Member access: sprint creation is restricted.")

    st.divider()
    sdf=load_sprints(pid=proj_id)
    if sdf.empty: st.info("No sprints yet."); st.markdown('</div>',unsafe_allow_html=True); st.stop()

    # Phase 1 Optimization: Load all tasks once and filter in memory
    all_project_tasks = load_tasks(project_id=proj_id, include_planning=True)
    has_active_sprint = (sdf["status"] == "Active").any()

    for _,s in sdf.iterrows():
        sid2=int(s["id"])
        if not all_project_tasks.empty:
            stask = all_project_tasks[all_project_tasks["sprint_id"] == sid2]
        else:
            stask = pd.DataFrame()
        dp=int(s["completed_points"]); pp=int(s["planned_points"]); pct=dp/max(pp,1)*100
        icon={"Planning":"🔵","Active":"🟢","Completed":"✅"}.get(s["status"],"⚪")
        with st.expander(f"{icon} **{s['name']}**  ·  {s['status']}  ·  {dp}/{pp} pts  ·  {s['start_date']} → {s['end_date']}",expanded=(s["status"]=="Active")):
            st.markdown(f"""
            <div style="background:rgba(255,255,255,0.02); border:1px solid rgba(255,255,255,0.05); border-radius:8px; padding:16px; margin-bottom:16px;">
                <div style="display:grid; grid-template-columns: 2fr 1fr; gap:16px;">
                    <div>
                        <span style="color:#8b949e; font-size:11px; text-transform:uppercase; letter-spacing:0.5px; font-weight:600;">Sprint Goal</span>
                        <div style="color:#e6edf3; font-weight:500; font-size:14px; margin-top:4px;">{s.get('goal') or '—'}</div>
                    </div>
                    <div style="text-align:right;">
                        <span style="color:#8b949e; font-size:11px; text-transform:uppercase; letter-spacing:0.5px; font-weight:600;">Duration</span>
                        <div style="color:#e6edf3; font-weight:500; font-size:13.5px; margin-top:4px;">📅 {s['start_date']} to {s['end_date']}</div>
                    </div>
                </div>
            </div>
            """, unsafe_allow_html=True)

            el,er=st.columns([2,1])
            with el:
                st.markdown(f"**Goal:** {s.get('goal') or '—'}"); st.progress(min(pct/100,1.0),text=f"{pct:.0f}% · {len(stask)} tasks")
                if st.session_state.get("role") == "Manager":
                    with st.form(f"es_{sid2}"):
                        c1,c2=st.columns(2)
                        en=c1.text_input("Name",value=s["name"]); eg=c1.text_input("Goal",value=s.get("goal") or "")
                        esd=c2.date_input("Start",value=pd.to_datetime(s["start_date"]).date()); eed=c2.date_input("End",value=pd.to_datetime(s["end_date"]).date())
                        ep=c2.number_input("Planned Pts",0,300,int(s["planned_points"] or 0))
                        if st.form_submit_button("💾 Save"):
                              exe("UPDATE sprints SET name=%s,goal=%s,start_date=%s,end_date=%s,planned_points=%s WHERE id=%s",(en,eg,str(esd),str(eed),ep,sid2)); st.success("Updated!"); st.rerun()
                else:
                    st.markdown(
                        f"""
                        <div style="background:#161b22;border:1px solid #30363d;border-radius:6px;padding:12px;margin-top:10px">
                            <table style="width:100%;border-collapse:collapse;color:#c9d1d9;font-size:13px">
                                <tr>
                                    <td style="padding:4px 0;color:#8b949e">Start Date:</td>
                                    <td style="padding:4px 0;font-weight:500">{s['start_date']}</td>
                                </tr>
                                <tr>
                                    <td style="padding:4px 0;color:#8b949e">End Date:</td>
                                    <td style="padding:4px 0;font-weight:500">{s['end_date']}</td>
                                </tr>
                                <tr>
                                    <td style="padding:4px 0;color:#8b949e">Planned Points:</td>
                                    <td style="padding:4px 0;font-weight:500">{s['planned_points']} pts</td>
                                </tr>
                            </table>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )

            with er:
                st.markdown("**Actions**")
                if st.session_state.get("role") == "Manager":
                    if s["status"]=="Planning":
                        if has_active_sprint: st.warning("Another sprint is active.")
                        elif st.button("▶️ Start Sprint",key=f"st_{sid2}",type="primary"):
                            exe("UPDATE sprints SET status='Active' WHERE id=%s",(sid2,)); log_activity(proj_id,"User","started sprint",s["name"]); st.success("Started!"); st.rerun()
                    if s["status"]=="Active":
                        if st.button("✅ Complete Sprint",key=f"cp_{sid2}",type="primary"):
                            exe("UPDATE tasks SET sprint_id=NULL WHERE sprint_id=%s AND status!='Done'",(sid2,))
                            exe("UPDATE sprints SET status='Completed' WHERE id=%s",(sid2,))
                            tl=stask.to_dict("records"); dc=sum(1 for t in tl if t["status"]=="Done"); bc=sum(1 for t in tl if t["status"]=="Blocked")
                            exe("INSERT INTO sprint_metrics(sprint_id,velocity,completion_rate,avg_cycle_time,blockers_count,on_time_tasks,late_tasks)VALUES(%s,%s,%s,%s,%s,%s,%s) ON CONFLICT (sprint_id) DO UPDATE SET velocity = EXCLUDED.velocity, completion_rate = EXCLUDED.completion_rate, avg_cycle_time = EXCLUDED.avg_cycle_time, blockers_count = EXCLUDED.blockers_count, on_time_tasks = EXCLUDED.on_time_tasks, late_tasks = EXCLUDED.late_tasks",
                                (sid2,dp,round(dp/max(pp,1)*100,1),2.5,bc,dc,len(tl)-dc))
                            log_activity(proj_id,"User","completed sprint",s["name"]); st.success("Completed!"); st.rerun()
                    if s["status"]=="Planning" and st.button("🗑️ Delete",key=f"dl_{sid2}"):
                        exe("DELETE FROM sprint_metrics WHERE sprint_id=%s",(sid2,)); exe("UPDATE tasks SET sprint_id=NULL WHERE sprint_id=%s",(sid2,)); exe("DELETE FROM sprints WHERE id=%s",(sid2,)); st.success("Deleted."); st.rerun()
                else:
                    st.caption("Member access: sprint actions are restricted.")
            st.divider(); st.markdown("**📥 Add Backlog Tasks**")
            if not all_project_tasks.empty:
                backlog = all_project_tasks[all_project_tasks["sprint_id"].isna() | (all_project_tasks["sprint_id"] == 0) | (all_project_tasks["sprint_id"] == None)]
            else:
                backlog = pd.DataFrame()
            if backlog.empty: st.caption("No unassigned tasks.")
            else:
                if st.session_state.get("role") == "Manager":
                    with st.form(f"at_{sid2}"):
                        sel_t2=st.multiselect("Select tasks",backlog["title"].tolist())
                        if st.form_submit_button("📥 Add to Sprint"):
                            for tit in sel_t2:
                                tr2=backlog[backlog["title"]==tit].iloc[0]; exe("UPDATE tasks SET sprint_id=%s WHERE id=%s",(sid2,int(tr2["id"]))); exe("UPDATE sprints SET planned_points=planned_points+%s WHERE id=%s",(int(tr2["story_points"]),sid2))
                            st.success(f"Added {len(sel_t2)}!"); st.rerun()
                else:
                    st.caption("Member access: backlog-to-sprint assignment is restricted.")
            if not stask.empty:
                sc2b=["title","assignee_name","priority","status","story_points"]
                st.dataframe(stask[[c for c in sc2b if c in stask.columns]].fillna("—"),use_container_width=True,hide_index=True)
    st.markdown('</div>',unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  🗒️ TASK CREATION
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Task Creation":
    st.markdown('<div class="page-wrap">', unsafe_allow_html=True)
    team=load_team(proj_id); sdf=load_sprints(pid=proj_id); opts2=member_opts(team)

    if st.session_state.get("role") == "Manager":
        with st.expander("➕ Create New Task",expanded=True):
            with st.form("ct", clear_on_submit=True):
                st.markdown("#### New Task")
                r1,r2=st.columns(2)
                t_tit=r1.text_input("Title *",placeholder="e.g. Implement login API")
                t_des=r1.text_area("Description",height=70,placeholder="What needs to be done?")
                t_typ=r1.selectbox("Issue Type",["Story","Task","Bug","Subtask","Epic"])
                t_tag=r1.text_input("Labels / Tags",placeholder="backend, api")
                t_spr=r2.selectbox("Sprint",[None]+(sdf["id"].tolist() if not sdf.empty else []),format_func=lambda x:"— Backlog —" if x is None else sdf[sdf["id"]==x]["name"].values[0])
                t_as=r2.selectbox("Assignee",list(opts2.keys()),format_func=lambda x:opts2[x])
                t_pr=r2.selectbox("Priority",["Medium","Low","High","Critical"])
                t_sp=r2.selectbox("Story Points",[1,2,3,5,8,13],index=2)
                t_eh=r2.number_input("Estimated Hours",0.5,200.0,float(t_sp*2),step=0.5)
                t_bl=r2.text_input("Blocker Note (optional)","")
                if st.form_submit_button("➕ Create Task",type="primary"):
                    if not t_tit.strip(): st.error("Title required.")
                    else:
                        vel3=14.0
                        if t_as:
                            row3=team[team["id"]==t_as]
                            if not row3.empty: vel3=float(row3["velocity_avg"].values[0])
                        cap=40
                        if t_spr and not sdf.empty:
                            sr3=sdf[sdf["id"]==t_spr]
                            if not sr3.empty: cap=int(sr3["planned_points"].values[0])
                        pred=get_predictor_for_project(proj_id)[0].predict(t_sp,t_eh,t_pr,vel3,cap)
                        init="Blocked" if t_bl.strip() else "Todo"
                        nid=exe("INSERT INTO tasks(sprint_id,project_id,title,description,assignee_id,priority,status,issue_type,story_points,estimated_hours,tags,blocker_note)VALUES(%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                                (t_spr,proj_id,t_tit.strip(),t_des.strip(),t_as,t_pr,init,t_typ,t_sp,t_eh,t_tag.strip(),t_bl.strip() or None))
                        if t_spr: exe("UPDATE sprints SET planned_points=planned_points+%s WHERE id=%s",(t_sp,t_spr))
                        actor=opts2.get(t_as,"User").split(" (")[0]
                        log_activity(proj_id,actor,"created",t_tit.strip(),"status","Todo",nid,t_spr,t_tit.strip())
                        c1,c2,c3=st.columns(3)
                        c1.metric("AI Predicted",f"{pred}h"); c2.metric("Your Estimate",f"{t_eh}h"); c3.metric("Difference",f"{pred-t_eh:+.1f}h")
                        risk4=risk_det.assess_task_risk({"priority":t_pr,"status":init,"story_points":t_sp,"estimated_hours":t_eh,"assignee_id":t_as},vel3,7)
                        ri4={"low":"🟢","medium":"🟡","high":"🔴"}[risk4["level"]]
                        st.info(f"AI Risk: {ri4} **{risk4['level'].upper()}**" + (f"  — {risk4['reasons'][0]}" if risk4["reasons"] else ""))
                        st.success(f"Task '{t_tit}' created (ID {nid})"); st.rerun()
    else:
        st.caption("Member access: task creation is restricted.")

    st.divider()
    fc1,fc2,fc3,fc4=st.columns(4)
    f_s2=fc1.selectbox("Sprint",["All"]+(sdf["name"].tolist() if not sdf.empty else []),key="bf1")
    f_st2=fc2.selectbox("Status",["All","Todo","In Progress","Done","Blocked"],key="bf2")
    f_p2=fc3.selectbox("Priority",["All","Critical","High","Medium","Low"],key="bf3")
    f_a2=fc4.selectbox("Assignee",["All"]+(team["name"].tolist() if not team.empty else []),key="bf4")

    all_t2=load_tasks(project_id=proj_id)
    if not all_t2.empty:
        if f_s2!="All" and not sdf.empty:
            sid3=int(sdf[sdf["name"]==f_s2]["id"].values[0]); all_t2=all_t2[all_t2["sprint_id"]==sid3]
        if f_st2!="All": all_t2=all_t2[all_t2["status"]==f_st2]
        if f_p2!="All":  all_t2=all_t2[all_t2["priority"]==f_p2]
        if f_a2!="All":  all_t2=all_t2[all_t2["assignee_name"]==f_a2]

    if all_t2.empty: st.caption("No tasks.")
    else:
        st.caption(f"{len(all_t2)} task(s)")
        PRI_ICON = {"Critical":"🔴","High":"🟡","Medium":"🔵","Low":"⚪"}
        STS_ICON = {"Done":"✅","In Progress":"⏳","Todo":"○","Blocked":"⛔"}
        for _,t in all_t2.iterrows():
            pri_txt = PRI_ICON.get(t['priority'],'') + ' ' + t['priority']
            sts_txt = STS_ICON.get(t['status'],'') + ' ' + t['status']
            av_txt  = t.get('assignee_name') or 'Unassigned'
            label   = f"{pri_txt}  {t['title']}  ·  {sts_txt}  ·  {t['story_points']}pt  ·  👤 {av_txt}"
            with st.expander(label, expanded=False):
                dc1,dc2=st.columns(2)
                with dc1:
                    st.markdown(f"**Desc:** {t.get('description') or '—'}")
                    st.markdown(f"**Tags:** `{t.get('tags') or '—'}` · **Type:** `{t.get('issue_type','Task')}`")
                    st.markdown(f"**Est:** {t.get('estimated_hours') or '—'}h · **Actual:** {t.get('actual_hours') or '—'}h")
                    if t.get("blocker_note"): st.error(f"⛔ {t['blocker_note']}")
                with dc2:
                    if st.session_state.get("role") == "Manager":
                        with st.form(f"te_{t['id']}"):
                            e_t=st.text_input("Title",value=t["title"],key=f"ett_{t['id']}")
                            e_s=st.selectbox("Status",["Todo","In Progress","Done","Blocked"],index=["Todo","In Progress","Done","Blocked"].index(t["status"]),key=f"ets_{t['id']}")
                            e_p=st.selectbox("Priority",["Low","Medium","High","Critical"],index=["Low","Medium","High","Critical"].index(t["priority"]),key=f"etp_{t['id']}")
                            e_a=st.selectbox("Assignee",list(opts2.keys()),format_func=lambda x:opts2[x],index=list(opts2.keys()).index(t["assignee_id"]) if t["assignee_id"] in list(opts2.keys()) else 0,key=f"eta_{t['id']}")
                            e_sp2=st.selectbox("Pts",[1,2,3,5,8,13],index=[1,2,3,5,8,13].index(int(t["story_points"])) if int(t["story_points"]) in [1,2,3,5,8,13] else 2,key=f"etsp_{t['id']}")
                            e_eh=st.number_input("Est.h",0.5,200.0,float(t.get("estimated_hours") or 2.0),step=0.5,key=f"eteh_{t['id']}")
                            e_ah=st.number_input("Act.h",0.0,200.0,float(t.get("actual_hours") or 0.0),step=0.5,key=f"etah_{t['id']}")
                            e_bl=st.text_input("Blocker",value=t.get("blocker_note") or "",key=f"etbl_{t['id']}")
                            sp_map={None:"— Backlog —"};sp_map.update({r["id"]:r["name"] for _,r in sdf.iterrows()} if not sdf.empty else {})
                            cur_spr=t["sprint_id"] if t["sprint_id"] in list(sp_map.keys()) else None
                            e_spr2=st.selectbox("Sprint",list(sp_map.keys()),format_func=lambda x:sp_map[x],index=list(sp_map.keys()).index(cur_spr),key=f"etspr_{t['id']}")
                            sv,dl=st.columns(2)
                            if sv.form_submit_button("💾",type="primary"):
                                old_pts=int(t["story_points"])
                                exe("UPDATE tasks SET title=%s,status=%s,priority=%s,assignee_id=%s,story_points=%s,estimated_hours=%s,actual_hours=%s,blocker_note=%s,sprint_id=%s,updated_at=CURRENT_TIMESTAMP WHERE id=%s",
                                    (e_t,e_s,e_p,e_a,e_sp2,e_eh,e_ah or None,e_bl or None,e_spr2,int(t["id"])))
                                old_sprint_id = t.get("sprint_id")

                                if pd.notna(old_sprint_id) and old_sprint_id != e_spr2:
                                    exe(
                                        "UPDATE sprints SET planned_points=GREATEST(0,planned_points-%s) WHERE id=%s",
                                        (old_pts, int(old_sprint_id))
                                    )
                                if pd.notna(e_spr2) and e_spr2 != t.get("sprint_id"): exe("UPDATE sprints SET planned_points=planned_points+%s WHERE id=%s",(e_sp2,e_spr2))
                                if e_s=="Done" and t["status"]!="Done" and pd.notna(e_spr2): exe("UPDATE sprints SET completed_points=completed_points+%s WHERE id=%s",(e_sp2,e_spr2))
                                elif t["status"]=="Done" and e_s!="Done" and pd.notna(e_spr2): exe("UPDATE sprints SET completed_points=GREATEST(0,completed_points-%s) WHERE id=%s",(e_sp2,e_spr2))
                                log_activity(proj_id,opts2.get(e_a, "User").split(" (")[0],"updated",e_t,"status",t["status"],e_s,int(t["id"]),e_spr2)
                                st.success("Saved!"); st.rerun()
                            if dl.form_submit_button("🗑️"):
                                if t.get("sprint_id"):
                                    exe("UPDATE sprints SET planned_points=GREATEST(0,planned_points-%s) WHERE id=%s",(int(t["story_points"]),int(t["sprint_id"])))
                                    if t["status"]=="Done": exe("UPDATE sprints SET completed_points=GREATEST(0,completed_points-%s) WHERE id=%s",(int(t["story_points"]),int(t["sprint_id"])))
                                exe("DELETE FROM tasks WHERE id=%s",(int(t["id"]),)); st.success("Deleted."); st.rerun()
                    else:
                        sp_map={None:"— Backlog —"};sp_map.update({r["id"]:r["name"] for _,r in sdf.iterrows()} if not sdf.empty else {})
                        cur_spr=t["sprint_id"] if t["sprint_id"] in list(sp_map.keys()) else None
                        spr_name = sp_map.get(cur_spr, "— Backlog —")
                        st.markdown(
                            f"""
                            <div style="background:#161b22;border:1px solid #30363d;border-radius:6px;padding:12px">
                                <table style="width:100%;border-collapse:collapse;color:#c9d1d9;font-size:13px">
                                    <tr>
                                        <td style="padding:4px 0;color:#8b949e">Sprint:</td>
                                        <td style="padding:4px 0;font-weight:500">{spr_name}</td>
                                    </tr>
                                    <tr>
                                        <td style="padding:4px 0;color:#8b949e">Status:</td>
                                        <td style="padding:4px 0;font-weight:500">{t['status']}</td>
                                    </tr>
                                    <tr>
                                        <td style="padding:4px 0;color:#8b949e">Priority:</td>
                                        <td style="padding:4px 0;font-weight:500">{t['priority']}</td>
                                    </tr>
                                    <tr>
                                        <td style="padding:4px 0;color:#8b949e">Assignee:</td>
                                        <td style="padding:4px 0;font-weight:500">{av_txt}</td>
                                    </tr>
                                    <tr>
                                        <td style="padding:4px 0;color:#8b949e">Story Points:</td>
                                        <td style="padding:4px 0;font-weight:500">{t['story_points']} pt</td>
                                    </tr>
                                </table>
                            </div>
                            """,
                            unsafe_allow_html=True
                        )
    st.markdown('</div>',unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  👥 TEAM
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Team":
    st.markdown('<div class="page-wrap">', unsafe_allow_html=True)
    team_df=load_team(proj_id)
    all_tasks=load_tasks(project_id=proj_id)
    act=load_active(proj_id)
    tasks_df=all_tasks[all_tasks["sprint_id"] == int(act["id"])] if (act and not all_tasks.empty) else pd.DataFrame()
    ROLES=["Frontend Developer","Backend Developer","Full Stack Developer","ML Engineer","Database Engineer","Data Scientist","DevOps Engineer","QA Engineer","Scrum Master","Product Owner","Tech Lead"]
    AVCOLS=["#ec4899","#3b82f6","#10b981","#f59e0b","#8b5cf6","#06b6d4","#f43f5e","#84cc16"]

    if st.session_state.get("role") == "Manager":
        with st.expander("➕ Add Team Member",expanded=team_df.empty):
            with st.form("am", clear_on_submit=True):
                mc1,mc2,mc3,mc4=st.columns(4)
                mn=mc1.text_input("Full Name *"); me=mc2.text_input("Email *"); mr=mc3.selectbox("Role",ROLES); mv=mc4.number_input("Velocity",1.0,60.0,12.0,step=0.5)
                if st.form_submit_button("➕ Add",type="primary"):
                    if not mn.strip(): st.error("Name required.")
                    elif not me.strip(): st.error("Email required.")
                    elif not team_df.empty and mn.strip() in team_df["name"].tolist(): st.error("Name already exists.")
                    elif not team_df.empty and me.strip() in team_df["email"].fillna("").tolist(): st.error("Email already exists.")
                    else:
                        col2=AVCOLS[len(team_df)%len(AVCOLS)]
                        exe("INSERT INTO team_members(name,role,velocity_avg,project_id,avatar_color,email)VALUES(%s,%s,%s,%s,%s,%s)",(mn.strip(),mr,mv,proj_id,col2,me.strip()))
                        log_activity(proj_id,mn.strip(),"joined team"); st.success(f"✅ {mn} added!"); st.rerun()

    st.divider()
    if team_df.empty: st.info("No team members yet."); st.markdown('</div>',unsafe_allow_html=True); st.stop()

    rows=[]
    for _,m in team_df.iterrows():
        mt=tasks_df[tasks_df["assignee_id"]==m["id"]] if not tasks_df.empty else pd.DataFrame()
        rows.append({"Name":m["name"],"Email":m.get("email") or "—","Role":m["role"],"Velocity":m["velocity_avg"],"Sprint Tasks":len(mt),
                     "Done":len(mt[mt["status"]=="Done"]) if not mt.empty else 0,"Blocked":len(mt[mt["status"]=="Blocked"]) if not mt.empty else 0})
    st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)
    st.divider()

    for i,(_,m) in enumerate(team_df.iterrows()):
        avc=m.get("avatar_color","#3b82f6"); mt=tasks_df[tasks_df["assignee_id"]==m["id"]] if not tasks_df.empty else pd.DataFrame()
        with st.expander(f"  **{m['name']}**  ·  {m['role']}  ·  ⚡ {m['velocity_avg']}pt",expanded=False):
            ec1,ec2=st.columns([1.5,1])
            with ec1:
                if not mt.empty:
                    sc3=["title","priority","status","story_points"]
                    st.dataframe(mt[[c for c in sc3 if c in mt.columns]].fillna("—"),use_container_width=True,hide_index=True)
                else: st.caption("No tasks this sprint.")
            with ec2:
                if st.session_state.get("role") == "Manager":
                    with st.form(f"em_{m['id']}"):
                        en2=st.text_input("Name",value=m["name"])
                        ee2=st.text_input("Email",value=m.get("email") or "")
                        er2=st.selectbox("Role",ROLES,index=ROLES.index(m["role"]) if m["role"] in ROLES else 0)
                        ev2=st.number_input("Velocity",1.0,60.0,float(m["velocity_avg"]),step=0.5)
                        sv3,rm3=st.columns(2)
                        if sv3.form_submit_button("💾 Save",type="primary"):
                            exe("UPDATE team_members SET name=%s,email=%s,role=%s,velocity_avg=%s WHERE id=%s",(en2,ee2.strip(),er2,ev2,int(m["id"]))); st.success("Saved!"); st.rerun()
                        if rm3.form_submit_button("🗑️ Remove"):
                            exe("UPDATE tasks SET assignee_id=NULL WHERE assignee_id=%s",(int(m["id"]),)); exe("DELETE FROM team_members WHERE id=%s",(int(m["id"]),)); st.success("Removed."); st.rerun()
                else:
                    st.markdown(
                        f"""
                        <div style="background:#161b22;border:1px solid #30363d;border-radius:6px;padding:12px">
                            <table style="width:100%;border-collapse:collapse;color:#c9d1d9;font-size:13px">
                                <tr>
                                    <td style="padding:4px 0;color:#8b949e">Email:</td>
                                    <td style="padding:4px 0;font-weight:500">{m.get('email') or '—'}</td>
                                </tr>
                                <tr>
                                    <td style="padding:4px 0;color:#8b949e">Role:</td>
                                    <td style="padding:4px 0;font-weight:500">{m['role']}</td>
                                </tr>
                                <tr>
                                    <td style="padding:4px 0;color:#8b949e">Velocity:</td>
                                    <td style="padding:4px 0;font-weight:500">{m['velocity_avg']} pt</td>
                                </tr>
                            </table>
                        </div>
                        """,
                        unsafe_allow_html=True
                    )
    st.markdown('</div>',unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  🔮 AI INSIGHTS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "AI Insights":
    st.markdown('<div class="page-wrap">', unsafe_allow_html=True)
    st.markdown("### 🔮 AI Insights & Predictions")
    st.caption("Auto-Blocker Detection · Risk Scoring · Recommendations")

    act=load_active(proj_id); team=load_team(proj_id)
    tab1,tab2,tab3=st.tabs(["🤖 Auto Detection","🔍 Risk","💡 Recommendations"])

    with tab1:
        st.subheader("🤖 Automatic Blocker & Delay Detection")
        st.markdown("""
        <div class="rc rc-act">
          <strong>How it works — No manual input needed</strong><br>
          <span style="color:#8b949e">The AI continuously monitors all tasks and automatically flags them based on:
          stagnation (no updates in 2+ days), hour overruns (actual > 130% of estimate),
          deadline proximity (not started with &lt;3 days left),
          high-priority tasks past sprint midpoint, and velocity mismatches.</span>
        </div>""", unsafe_allow_html=True)

        if not act:
            st.info("No active sprint to analyze.")
        else:
            # Phase 2 Optimization: Use cached sprint scan
            scan5 = get_sprint_scan_cached(int(act["id"]))

            c1,c2,c3,c4=st.columns(4)
            c1.metric("Auto-Flagged",  scan5["auto_blocked_count"], help="Tasks AI detected as effectively blocked")
            c2.metric("Delayed",       scan5["delayed_count"],      help="Tasks predicted to miss deadline")
            c3.metric("At Risk",       scan5["at_risk_count"],      help="Tasks showing warning signs")
            c4.metric("Days Remaining",scan5["remaining_days"])

            st.divider()
            flagged=[t for t in scan5["tasks"] if t["auto_detection"]["delay_risk"]!="none" or t["auto_detection"]["auto_blocked"]]
            if not flagged:
                st.markdown('<div class="rc rc-ok">✅ No issues auto-detected. All tasks look on track.</div>',unsafe_allow_html=True)
            else:
                for t in flagged:
                    det=t["auto_detection"]; ab=det["auto_blocked"]; dr=det["delay_risk"]
                    css="auto-blk" if ab or dr=="delayed" else "auto-risk"
                    bc="#f85149" if ab or dr=="delayed" else "#d29922"
                    av=t.get("assignee_name") or "Unassigned"; avc=t.get("avatar_color") or "#8b949e"
                    reasons_html="".join(f'<li style="color:#c9d1d9;font-size:12px">{r}</li>' for r in det["reasons"])
                    sugg_html="".join(f'<li style="color:#58a6ff;font-size:12px">💡 {s}</li>' for s in det["suggestions"])
                    delay_badge=f'<span style="color:#f85149;font-size:11px;font-weight:700">~{det["delay_days"]}d delay</span>' if det["delay_days"]>0 else ""
                    status_label = '⛔ AUTO-BLOCKED' if ab else '⏰ DELAYED' if dr=='delayed' else '⚠️ AT RISK'
                    st.markdown(
                        f'<div class="{css}" style="margin-bottom:10px">'
                        f'<div style="display:flex;justify-content:space-between;align-items:flex-start;margin-bottom:8px">'
                        f'<div><span style="color:{bc};font-weight:700;font-size:14px">{status_label}</span>'
                        f'<span style="color:#e6edf3;font-size:13px;font-weight:500;margin-left:8px">{t["title"]}</span></div>'
                        f'<div style="display:flex;gap:6px;align-items:center">{delay_badge}{av_html(av,avc,22)}'
                        f'<span style="color:#8b949e;font-size:11px">{av}</span>{pbadge(t["priority"])}</div>'
                        f'</div>'
                        f'<div style="display:grid;grid-template-columns:1fr 1fr;gap:8px">'
                        f'<div><div style="color:#8b949e;font-size:11px;font-weight:600;margin-bottom:3px">REASONS DETECTED</div>'
                        f'<ul style="margin:0;padding-left:14px">{reasons_html}</ul></div>'
                        f'<div><div style="color:#8b949e;font-size:11px;font-weight:600;margin-bottom:3px">AI SUGGESTIONS</div>'
                        f'<ul style="margin:0;padding-left:14px">{sugg_html}</ul></div>'
                        f'</div></div>',
                        unsafe_allow_html=True
                    )

    with tab2:
        st.subheader("🔍 Risk Analyzer")
        if not act: st.warning("No active sprint.")
        else:
            tdf6=load_tasks(sprint_id=act["id"]); tasks6=tdf6.to_dict("records")
            sr6=risk_det.assess_sprint_risk(act,tasks6)
            rc6={"low":"#3fb950","medium":"#d29922","high":"#f85149"}[sr6["level"]]
            st.markdown(f'<div style="background:{rc6}15;border:1px solid {rc6}40;border-radius:8px;padding:14px;margin-bottom:14px">'
                        f'<span style="color:{rc6};font-size:18px;font-weight:700">Sprint Risk: {sr6["level"].upper()}</span>'
                        f'<span style="color:#8b949e;font-size:13px"> (score {sr6["score"]})</span>'
                        f'{"".join(f"<br><span style=color:#8b949e;font-size:12px>⚠️ {r}</span>" for r in sr6["reasons"])}'
                        f'</div>',unsafe_allow_html=True)
            rrows=[]
            for t in tasks6:
                tr7=risk_det.assess_task_risk(t,t.get("velocity_avg",12),7)
                rrows.append({"Task":t["title"],"Assignee":t.get("assignee_name","—"),"Priority":t["priority"],"Status":t["status"],"Risk":tr7["level"].upper(),"Score":tr7["score"],"Reason":tr7["reasons"][0] if tr7["reasons"] else "—"})
            if rrows: st.dataframe(pd.DataFrame(rrows).sort_values("Score",ascending=False),use_container_width=True,hide_index=True)

    with tab3:
        st.subheader("💡 Recommendations")
        if not act: st.info("No active sprint.")
        else:
            tdf7=load_tasks(sprint_id=act["id"]); tasks7=tdf7.to_dict("records")
            rb_insights = generate_insights(tasks7)
            rb_blockers, rb_warnings = detect_blockers(tasks7)
            if rb_insights:
                st.markdown("**Rule-Based Insights**")
                for msg in rb_insights[:2]:
                    st.markdown(f'<div class="rc rc-info">💡 {msg}</div>', unsafe_allow_html=True)
            if rb_warnings:
                st.markdown("**Blocker Detection Warnings**")
                for warn in rb_warnings[:2]:
                    st.markdown(f'<div class="rc rc-warn">⚠️ {warn}</div>', unsafe_allow_html=True)

            mets6=load_metrics(proj_id); vl6=mets6["velocity"].tolist() if not mets6.empty else [30,35,38,40]
            sr7=risk_det.assess_sprint_risk(act,tasks7)
            # Use a lightweight np.mean heuristic instead of calling expensive ML forecasting
            vfc7 = {"forecast": int(np.mean(vl6)) if vl6 else 30, "trend": "stable", "confidence": 0.8}
            recs7=rec_eng.generate(sr7,tasks7,vfc7,team.to_dict("records"))
            rc_map2={"warning":"rc-warn","caution":"rc-caut","action":"rc-act","info":"rc-info","success":"rc-ok"}
            st.markdown("**🔔 Recommendations**")
            for r in recs7[:2]:
                st.markdown(f'<div class="rc {rc_map2.get(r["type"],"rc-info")}"><strong style="color:#e6edf3">{r["icon"]} {r["title"]}</strong><br><span style="color:#8b949e">{r["body"]}</span></div>',unsafe_allow_html=True)
    st.markdown('</div>',unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  📄 REPORTS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Reports":
    st.markdown('<div class="page-wrap">', unsafe_allow_html=True)
    
    current_user = st.session_state.get("user")
    if (
        "last_report_user" not in st.session_state
        or st.session_state.last_report_user != current_user
    ):
        st.session_state.last_report_user = current_user
        st.session_state.reports_state = {
            "sprint_name": None,
            "generated": False
        }

    sdf7=load_sprints(pid=proj_id)
    if sdf7.empty: st.info("No sprints yet."); st.markdown('</div>',unsafe_allow_html=True); st.stop()

    sel7=st.selectbox("Select Sprint",sdf7["name"].tolist())

    if "reports_state" not in st.session_state:
        st.session_state.reports_state = {"sprint_name": None, "generated": False}

    if sel7 != st.session_state.reports_state["sprint_name"]:
        st.session_state.reports_state = {"sprint_name": sel7, "generated": False}

    if not st.session_state.reports_state["generated"]:
        st.markdown(f"""
        <div style="background:#161b22;border:1px solid #30363d;border-radius:8px;padding:32px;text-align:center;margin-bottom:16px">
          <div style="font-size:40px;margin-bottom:12px">📄</div>
          <div style="color:#e6edf3;font-size:16px;font-weight:600;margin-bottom:6px">Sprint Performance Report</div>
          <div style="color:#8b949e;font-size:13px;margin-bottom:20px">
            Generate a detailed performance report for <strong>{sel7}</strong>. This compiles sprint metrics, health breakdown, AI risk scores, team workloads, and exportable logs.
          </div>
        </div>
        """, unsafe_allow_html=True)
        if st.button("📄 View Sprint Report", type="primary", use_container_width=True):
            st.session_state.reports_state["generated"] = True
            st.rerun()
        st.markdown('</div>', unsafe_allow_html=True)
        st.stop()

    srow7=sdf7[sdf7["name"]==sel7].iloc[0]; sid7=int(srow7["id"])
    tdf7=load_tasks(sprint_id=sid7); tasks7=tdf7.to_dict("records")
    hlth7=compute_sprint_health(tasks7,srow7.to_dict()); sr7b=risk_det.assess_sprint_risk(srow7.to_dict(),tasks7)
    pp7=int(srow7.get("planned_points",0)); cp7=int(srow7.get("completed_points",0))
    gc7={"A":"#3fb950","B":"#d29922","C":"#f0883e","D":"#f85149","F":"#f85149"}.get(hlth7["grade"],"#8b949e")
    rc7={"low":"#3fb950","medium":"#d29922","high":"#f85149"}[sr7b["level"]]

    stats=[("Story Points",f"{cp7}/{pp7}","#e6edf3"),("Health Score",f"{hlth7['score']}",gc7),
           ("Risk",sr7b["level"].upper(),rc7),("Tasks",len(tasks7),"#e6edf3"),
           ("Done",sum(1 for t in tasks7 if t["status"]=="Done"),"#3fb950"),
           ("Blocked",sum(1 for t in tasks7 if t["status"]=="Blocked"),"#f85149")]
    st.markdown(f"""
    <div class="card" style="margin-bottom:16px">
      <div style="font-size:18px;font-weight:700;color:#e6edf3;margin-bottom:4px">{sel7}</div>
      <div style="color:#8b949e;font-size:13px">📅 {srow7.get('start_date','—')} → {srow7.get('end_date','—')} · {srow7.get('goal') or '—'}</div>
      <div style="display:flex;gap:10px;margin-top:14px;flex-wrap:wrap">
        {''.join(f'<div style="background:#21262d;border-radius:6px;padding:10px 16px;text-align:center;min-width:80px"><div style="color:{c};font-size:20px;font-weight:700">{v}</div><div style="color:#8b949e;font-size:10px">{l}</div></div>' for l,v,c in stats)}
      </div>
    </div>""", unsafe_allow_html=True)

    # ── Excel Export (Managers only) ──
    if st.session_state.get("role") == "Manager":
        st.markdown('<div class="card">', unsafe_allow_html=True)
        st.markdown("#### 📥 Export Project Reports")
        st.caption("Download all project sprints, tickets, time logs, and member workloads as a single Excel spreadsheet.")
        
        try:
            import openpyxl
            openpyxl_installed = True
        except ImportError:
            openpyxl_installed = False
            
        if not openpyxl_installed:
            st.error("⚠️ The spreadsheet export engine (`openpyxl`) is missing. Please contact your system administrator to run `pip install openpyxl`.")
        else:
            def style_worksheet(ws):
                try:
                    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
                    from openpyxl.utils import get_column_letter
                    
                    try:
                        ws.views.sheetView[0].showGridLines = True
                    except Exception:
                        pass
                        
                    font_family = "Segoe UI"
                    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
                    cell_font = Font(name=font_family, size=10, color="000000")
                    
                    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
                    
                    green_fill = PatternFill(start_color="E6F4EA", end_color="E6F4EA", fill_type="solid")
                    green_font = Font(name=font_family, size=10, color="137333", bold=True)
                    
                    red_fill = PatternFill(start_color="FCE8E6", end_color="FCE8E6", fill_type="solid")
                    red_font = Font(name=font_family, size=10, color="C5221F", bold=True)
                    
                    yellow_fill = PatternFill(start_color="FEF7E0", end_color="FEF7E0", fill_type="solid")
                    yellow_font = Font(name=font_family, size=10, color="B06000", bold=True)
                    
                    blue_fill = PatternFill(start_color="E8F0FE", end_color="E8F0FE", fill_type="solid")
                    blue_font = Font(name=font_family, size=10, color="1A73E8", bold=True)
                    
                    thin_border_side = Side(style='thin', color='E5E7EB')
                    border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
                    
                    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    right_align = Alignment(horizontal='right', vertical='center', wrap_text=True)
                    
                    if ws.max_row >= 1:
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=1, column=col)
                            cell.font = header_font
                            cell.fill = header_fill
                            cell.alignment = center_align
                            cell.border = border
                            
                    for row in range(2, ws.max_row + 1):
                        for col in range(1, ws.max_column + 1):
                            cell = ws.cell(row=row, column=col)
                            cell.font = cell_font
                            cell.border = border
                            
                            val_str = str(cell.value or '').strip()
                            if cell.value is not None and isinstance(cell.value, (int, float)):
                                cell.alignment = right_align
                            else:
                                cell.alignment = left_align
                                
                            header_val = str(ws.cell(row=1, column=col).value or '').lower()
                            if header_val in ('status', 'priority', 'risk', 'issue_type', 'blocked_tasks'):
                                cell.alignment = center_align
                                val_lower = val_str.lower()
                                if val_lower in ('done', 'completed', 'active', 'low', 'task'):
                                    cell.fill = green_fill
                                    cell.font = green_font
                                elif val_lower in ('blocked', 'critical', 'high', 'delayed', 'bug'):
                                    cell.fill = red_fill
                                    cell.font = red_font
                                elif val_lower in ('in progress', 'medium', 'at_risk', 'story'):
                                    cell.fill = yellow_fill
                                    cell.font = yellow_font
                                elif val_lower in ('todo', 'backlog', 'qa'):
                                    cell.fill = blue_fill
                                    cell.font = blue_font
                                    
                    for col in ws.columns:
                        max_len = 0
                        col_letter = get_column_letter(col[0].column)
                        for cell in col[:100]:
                            val_str = str(cell.value or '')
                            lines = val_str.split('\n')
                            for line in lines:
                                if len(line) > max_len:
                                    max_len = len(line)
                        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)
                except Exception:
                    pass

            def generate_excel_data():
                buffer = io.BytesIO()
                with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                    # 1. Project Overview
                    proj_info = qry("SELECT name, description, created_by, created_at FROM projects WHERE id = %s", [proj_id])
                    if not proj_info.empty and 'created_at' in proj_info.columns:
                        proj_info['created_at'] = proj_info['created_at'].apply(lambda x: to_local_dt(x).strftime('%Y-%m-%d %H:%M:%S') if to_local_dt(x) else '')
                    proj_info.to_excel(writer, sheet_name='Project Overview', index=False)
                    
                    # 2. Sprints
                    sprints_info = qry("""
                        SELECT id, name, goal, start_date, end_date, status, planned_points, completed_points 
                        FROM sprints WHERE project_id = %s ORDER BY id
                    """, [proj_id])
                    if sprints_info.empty:
                        sprints_info = pd.DataFrame([{"Message": "No sprints recorded for this project."}])
                    sprints_info.to_excel(writer, sheet_name='Sprints', index=False)
                    
                    # 3. Tasks & Tickets
                    tasks_info = qry("""
                        SELECT t.id as task_id, t.title, t.description, tm.name as assignee_name, tm.email as assignee_email, 
                               s.name as sprint_name, t.priority, t.status, t.issue_type, t.story_points, t.estimated_hours, t.actual_hours, 
                               t.blocker_note, t.due_date, t.created_at, t.updated_at
                        FROM tasks t 
                        LEFT JOIN team_members tm ON t.assignee_id = tm.id 
                        LEFT JOIN sprints s ON t.sprint_id = s.id 
                        WHERE t.project_id = %s ORDER BY t.id
                    """, [proj_id])
                    if tasks_info.empty:
                        tasks_info = pd.DataFrame([{"Message": "No tasks recorded for this project."}])
                    else:
                        for col in ['created_at', 'updated_at']:
                            if col in tasks_info.columns:
                                 tasks_info[col] = tasks_info[col].apply(lambda x: to_local_dt(x).strftime('%Y-%m-%d %H:%M:%S') if to_local_dt(x) else '')
                    tasks_info.to_excel(writer, sheet_name='Tasks & Tickets', index=False)
                    
                    # 4. Daily Progress Logs
                    logs_info = qry("""
                        SELECT mc.id as log_id, mc.created_at as timestamp, tm.name as member_name, tm.email as member_email, 
                               s.name as sprint_name, t.title as task_title, mc.comment_text, mc.hours_logged
                        FROM member_comments mc
                        JOIN team_members tm ON mc.member_id = tm.id
                        LEFT JOIN sprints s ON mc.sprint_id = s.id
                        LEFT JOIN tasks t ON mc.task_id = t.id
                        WHERE mc.project_id = %s ORDER BY mc.id DESC
                    """, [proj_id])
                    if logs_info.empty:
                        logs_info = pd.DataFrame([{"Message": "No daily progress logs recorded for this project."}])
                    else:
                        if 'timestamp' in logs_info.columns:
                             logs_info['timestamp'] = logs_info['timestamp'].apply(lambda x: to_local_dt(x).strftime('%Y-%m-%d %H:%M:%S') if to_local_dt(x) else '')
                    logs_info.to_excel(writer, sheet_name='Daily Progress Logs', index=False)
                    
                    # 5. Team Workload
                    workload_info = qry("""
                        SELECT tm.id as member_id, tm.name, tm.email, tm.role, tm.velocity_avg,
                               (SELECT COUNT(*) FROM tasks t WHERE t.assignee_id = tm.id) as total_tasks,
                               (SELECT COUNT(*) FROM tasks t WHERE t.assignee_id = tm.id AND t.status = 'Done') as completed_tasks,
                               (SELECT COUNT(*) FROM tasks t WHERE t.assignee_id = tm.id AND t.status = 'Blocked') as blocked_tasks,
                               (SELECT COALESCE(SUM(t.actual_hours), 0) FROM tasks t WHERE t.assignee_id = tm.id) as total_hours_spent
                        FROM team_members tm 
                        WHERE tm.project_id = %s ORDER BY tm.id
                    """, [proj_id])
                    if workload_info.empty:
                        workload_info = pd.DataFrame([{"Message": "No team members registered for this project."}])
                    workload_info.to_excel(writer, sheet_name='Team Workload', index=False)
                    
                    # Style all worksheets
                    for name in writer.sheets:
                        style_worksheet(writer.sheets[name])
                        
                buffer.seek(0)
                return buffer.getvalue()
            
            clean_proj_name = "".join(c for c in proj_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
            filename = f"SprintAI_Export_{clean_proj_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            
            st.download_button(
                label="📥 Download Project Excel Report",
                data=generate_excel_data(),
                file_name=filename,
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
                type="primary"
            )
            
            st.markdown("---")
            st.markdown("#### 👤 Export Individual Team Member Report")
            st.caption("Select a team member to download their specific task assignments, daily logs, and performance metrics.")
            
            team_df = load_team(proj_id)
            if team_df.empty:
                st.info("No team members registered for this project yet.")
            else:
                member_opts = {int(row["id"]): f"{row['name']} ({row['role']})" for _, row in team_df.iterrows()}
                selected_m_id = st.selectbox("Select Team Member", list(member_opts.keys()), format_func=lambda x: member_opts[x])
                
                selected_m_name = team_df[team_df["id"] == selected_m_id].iloc[0]["name"]
                
                def generate_member_excel_data(m_id, m_name):
                    buffer = io.BytesIO()
                    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
                        # 1. Member Profile & Metrics
                        profile_info = qry("""
                            SELECT tm.name, tm.email, tm.role, tm.velocity_avg,
                                   (SELECT COUNT(*) FROM tasks t WHERE t.assignee_id = tm.id AND t.project_id = %s) as total_tasks,
                                   (SELECT COUNT(*) FROM tasks t WHERE t.assignee_id = tm.id AND t.project_id = %s AND t.status = 'Done') as completed_tasks,
                                   (SELECT COUNT(*) FROM tasks t WHERE t.assignee_id = tm.id AND t.project_id = %s AND t.status = 'Blocked') as blocked_tasks,
                                   (SELECT COALESCE(SUM(t.actual_hours), 0) FROM tasks t WHERE t.assignee_id = tm.id AND t.project_id = %s) as total_hours_spent
                            FROM team_members tm
                            WHERE tm.id = %s AND tm.project_id = %s
                        """, [proj_id, proj_id, proj_id, proj_id, m_id, proj_id])
                        profile_info.to_excel(writer, sheet_name='Member Profile', index=False)
                        
                        # 2. Assigned Tasks
                        member_tasks = qry("""
                            SELECT t.id as task_id, t.title, t.description, 
                                   s.name as sprint_name, t.priority, t.status, t.issue_type, t.story_points, t.estimated_hours, t.actual_hours, 
                                   t.blocker_note, t.due_date, t.created_at, t.updated_at
                            FROM tasks t 
                            LEFT JOIN sprints s ON t.sprint_id = s.id 
                            WHERE t.assignee_id = %s AND t.project_id = %s ORDER BY t.id
                        """, [m_id, proj_id])
                        if member_tasks.empty:
                            member_tasks = pd.DataFrame([{"Message": "No tasks assigned to this team member."}])
                        else:
                            for col in ['created_at', 'updated_at']:
                                if col in member_tasks.columns:
                                     member_tasks[col] = member_tasks[col].apply(lambda x: to_local_dt(x).strftime('%Y-%m-%d %H:%M:%S') if to_local_dt(x) else '')
                        member_tasks.to_excel(writer, sheet_name='Assigned Tasks', index=False)
                        
                        # 3. Daily Progress Logs
                        member_logs = qry("""
                            SELECT mc.id as log_id, mc.created_at as timestamp, 
                                   s.name as sprint_name, t.title as task_title, mc.comment_text, mc.hours_logged
                            FROM member_comments mc
                            LEFT JOIN sprints s ON mc.sprint_id = s.id
                            LEFT JOIN tasks t ON mc.task_id = t.id
                            WHERE mc.member_id = %s AND mc.project_id = %s ORDER BY mc.id DESC
                        """, [m_id, proj_id])
                        if member_logs.empty:
                            member_logs = pd.DataFrame([{"Message": "No daily progress logs recorded for this team member."}])
                        else:
                            if 'timestamp' in member_logs.columns:
                                 member_logs['timestamp'] = member_logs['timestamp'].apply(lambda x: to_local_dt(x).strftime('%Y-%m-%d %H:%M:%S') if to_local_dt(x) else '')
                        member_logs.to_excel(writer, sheet_name='Daily Progress Logs', index=False)
                        
                        # Style all worksheets
                        for name in writer.sheets:
                            style_worksheet(writer.sheets[name])
                            
                    buffer.seek(0)
                    return buffer.getvalue()
                
                clean_m_name = "".join(c for c in selected_m_name if c.isalnum() or c in (' ', '_', '-')).strip().replace(' ', '_')
                member_filename = f"SprintAI_Member_{clean_m_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                
                st.download_button(
                    label=f"📥 Download Report for {selected_m_name}",
                    data=generate_member_excel_data(selected_m_id, selected_m_name),
                    file_name=member_filename,
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
        st.markdown('</div>', unsafe_allow_html=True)

    r1,r2=st.columns(2)
    with r1:
        st.markdown("**🔍 Risk Factors**")
        if sr7b["reasons"]:
            for r in sr7b["reasons"]: st.markdown(f'<div class="rc rc-warn">⚠️ {r}</div>',unsafe_allow_html=True)
        else: st.markdown('<div class="rc rc-ok">✅ No major risks</div>',unsafe_allow_html=True)

        st.markdown("**🤖 Health Breakdown**")
        for comp,score in hlth7["breakdown"].items():
            maxs={"Completion Rate":40,"No Blockers":20,"Velocity Match":25,"Critical Tasks":15}.get(comp,10)
            pct8=score/maxs*100; col8="#3fb950" if pct8>=70 else "#d29922" if pct8>=40 else "#f85149"
            st.markdown(f"""
            <div style="margin-bottom:9px">
              <div style="display:flex;justify-content:space-between;margin-bottom:2px">
                <span style="color:#c9d1d9;font-size:12px">{comp}</span>
                <span style="color:{col8};font-size:12px;font-weight:600">{score}/{maxs}</span>
              </div>
              <div style="background:#21262d;border-radius:3px;height:5px">
                <div style="width:{pct8:.0f}%;background:{col8};height:5px;border-radius:3px"></div>
              </div>
            </div>""", unsafe_allow_html=True)
    with r2:
        st.markdown("**📊 Status Distribution**")
        if tasks7:
            st.dataframe(pd.DataFrame([{"Status":"Done","Count":sum(1 for t in tasks7 if t["status"]=="Done")},{"Status":"In Progress","Count":sum(1 for t in tasks7 if t["status"]=="In Progress")},{"Status":"Todo","Count":sum(1 for t in tasks7 if t["status"]=="Todo")},{"Status":"Blocked","Count":sum(1 for t in tasks7 if t["status"]=="Blocked")}]),use_container_width=True,hide_index=True)

    if not tdf7.empty:
        st.divider(); st.markdown("**All Tasks**")
        sc4=["title","assignee_name","priority","status","story_points","estimated_hours","actual_hours"]
        st.dataframe(tdf7[[c for c in sc4 if c in tdf7.columns]].fillna("—"),use_container_width=True,hide_index=True)
    st.caption("*Auto-generated · AI Sprint Manager *")
    st.markdown('</div>',unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
#  📝 DAILY PROGRESS
# ══════════════════════════════════════════════════════════════════════════════
elif page == "Daily Updates":
    st.markdown('<div class="page-wrap">', unsafe_allow_html=True)

    # Reset pagination limits on project switch
    if "last_project_id" not in st.session_state or st.session_state.last_project_id != proj_id:
        st.session_state.last_project_id = proj_id
        st.session_state.my_logs_limit = 20
        st.session_state.team_logs_limit = 20

    st.markdown("### 📝 Daily Progress Logs")
    st.caption("Log daily progress, update hours, and track team status updates.")
    
    role = st.session_state.get("role")
    user = st.session_state.get("user")
    
    # ── 1. Load active project details ──
    if not proj_id:
        st.warning("Please select or create a project first.")
        st.markdown('</div>', unsafe_allow_html=True)
        st.stop()
        
    sprints_df = load_sprints(pid=proj_id)
    active_sprint = load_active(proj_id)
    
    member_df = qry("SELECT * FROM team_members WHERE LOWER(email) = LOWER(%s) AND project_id = %s", [user, proj_id])
    
    # ── MEMBER VIEW ──
    if role == "Member":
        if member_df.empty:
            st.error("⚠️ You are not registered as a team member in this project. Please contact your manager.")
            st.markdown('</div>', unsafe_allow_html=True)
            st.stop()
            
        member_row = member_df.iloc[0]
        member_id = int(member_row["id"])
        member_name = member_row["name"]
        
        lcol, rcol = st.columns([1, 1.2])
        
        with lcol:
            st.markdown('<div class="card">', unsafe_allow_html=True)
            st.markdown(f"#### 📝 Log Today's Progress ({member_name})")
            
            with st.form("log_progress_form", clear_on_submit=True):
                sprint_opts = {None: "— No Sprint / Backlog —"}
                if not sprints_df.empty:
                    sprint_opts.update({int(row["id"]): f"{row['name']} ({row['status']})" for _, row in sprints_df.iterrows()})
                
                def_sprint_id = int(active_sprint["id"]) if active_sprint else None
                sprint_keys = list(sprint_opts.keys())
                sprint_idx = sprint_keys.index(def_sprint_id) if def_sprint_id in sprint_keys else 0
                
                sel_sprint_id = st.selectbox("Link to Sprint", sprint_keys, index=sprint_idx, format_func=lambda x: sprint_opts[x])
                
                task_query = "SELECT id, title, status FROM tasks WHERE assignee_id = %s AND project_id = %s"
                task_params = [member_id, proj_id]
                if sel_sprint_id:
                    task_query += " AND sprint_id = %s"
                    task_params.append(sel_sprint_id)
                task_query += " ORDER BY id DESC"
                
                tasks_df = qry(task_query, task_params)
                
                task_opts = {None: "— No Specific Task —"}
                if not tasks_df.empty:
                    task_opts.update({int(row["id"]): f"[{row['id']}] {row['title']} ({row['status']})" for _, row in tasks_df.iterrows()})
                    
                sel_task_id = st.selectbox("Link to Task (Optional)", list(task_opts.keys()), format_func=lambda x: task_opts[x])
                
                sprint_completed_and_locked = False
                if sel_sprint_id:
                    srow = sprints_df[sprints_df["id"] == sel_sprint_id]
                    if not srow.empty and srow.iloc[0]["status"] == "Completed":
                        end_date_str = srow.iloc[0].get("end_date")
                        if end_date_str:
                            try:
                                end_dt = pd.to_datetime(end_date_str)
                                time_diff = (datetime.now() - end_dt).total_seconds() / 3600
                                if time_diff > 48:
                                    sprint_completed_and_locked = True
                            except:
                                pass
                
                comment_val = st.text_area("What did you accomplish today? *", placeholder="Describe your progress, milestones, or blockers...")
                
                if sprint_completed_and_locked:
                    st.caption("🔒 Time logging is locked for this completed sprint (>48h ago). Only comments are allowed.")
                    hours_logged = 0.0
                else:
                    hours_logged = st.number_input("Time Spent Today (Hours)", min_value=0.0, max_value=24.0, value=0.0, step=0.5, help="This will be added to the task's total actual hours.")
                
                submit_log = st.form_submit_button("Submit Progress Log", type="primary", use_container_width=True)
                
                if submit_log:
                    if not comment_val.strip():
                        st.error("Progress comment is required.")
                    else:
                        if sel_task_id:
                            valid_df = qry("SELECT 1 FROM tasks WHERE id = %s AND project_id = %s AND assignee_id = %s", [sel_task_id, proj_id, member_id])
                            if valid_df.empty:
                                st.error("Mismatched task, project, or assignee.")
                                st.stop()
                                
                        new_comment_id = exe("""
                            INSERT INTO member_comments (project_id, member_id, sprint_id, task_id, comment_text, hours_logged)
                            VALUES (%s, %s, %s, %s, %s, %s)
                        """, (proj_id, member_id, sel_sprint_id, sel_task_id, comment_val.strip(), hours_logged))
                        
                        if sel_task_id and hours_logged > 0.0 and not sprint_completed_and_locked:
                            exe("""
                                UPDATE tasks 
                                SET actual_hours = COALESCE(actual_hours, 0) + %s, updated_at = CURRENT_TIMESTAMP 
                                WHERE id = %s
                            """, (hours_logged, sel_task_id))
                            
                            task_title_row = tasks_df[tasks_df["id"] == sel_task_id]
                            task_title = task_title_row.iloc[0]["title"] if not task_title_row.empty else "Task"
                            log_activity(proj_id, member_name, f"logged {hours_logged}h", task_title, "actual_hours", "", f"+{hours_logged}h", sel_task_id, sel_sprint_id)
                            
                        st.success("Daily progress log successfully submitted!")
                        st.rerun()
            st.markdown('</div>', unsafe_allow_html=True)
            
        with rcol:
            st.markdown("#### 🕒 My Recent Logs")
            
            if "my_logs_limit" not in st.session_state:
                st.session_state.my_logs_limit = 20

            count_df = qry("""
                SELECT COUNT(*) as cnt FROM member_comments 
                WHERE member_id = %s AND project_id = %s
            """, [member_id, proj_id])
            total_my_logs = int(count_df.iloc[0]["cnt"]) if not count_df.empty else 0

            my_logs = qry("""
                SELECT mc.*, s.name as sprint_name, t.title as task_title, t.id as t_id
                FROM member_comments mc
                LEFT JOIN sprints s ON mc.sprint_id = s.id
                LEFT JOIN tasks t ON mc.task_id = t.id
                WHERE mc.member_id = %s AND mc.project_id = %s
                ORDER BY mc.id DESC
                LIMIT %s
            """, [member_id, proj_id, st.session_state.my_logs_limit])
            
            if my_logs.empty:
                st.caption("You haven't logged any progress comments in this project yet.")
            else:
                for _, log in my_logs.iterrows():
                    log_id = int(log["id"])
                    created_at_dt = to_local_dt(log["created_at"])
                    if created_at_dt:
                        time_elapsed_hrs = (datetime.now(created_at_dt.tzinfo) - created_at_dt).total_seconds() / 3600
                        can_edit = time_elapsed_hrs <= 24.0
                        time_str = created_at_dt.strftime("%d %b %Y • %I:%M %p IST")
                    else:
                        can_edit = False
                        time_str = str(log["created_at"]) if pd.notna(log["created_at"]) else "Unknown Time"
                    task_info = f" · Task: **{log['task_title']}**" if log["task_title"] else ""
                    sprint_info = f" · Sprint: **{log['sprint_name']}**" if log["sprint_name"] else ""
                    hours_info = f" · Time: **{log['hours_logged']}h**" if float(log["hours_logged"]) > 0 else ""
                    
                    st.markdown(f"""
                    <div class="card" style="border-left: 3px solid #10b981;">
                        <div style="display:flex; justify-content:space-between; align-items:center; margin-bottom:8px;">
                            <span style="font-size:11px; color:#8b949e;">🕒 {time_str}{sprint_info}{task_info}{hours_info}</span>
                        </div>
                        <p style="color:#e6edf3; font-size:13px; margin:0;">{log['comment_text']}</p>
                    </div>
                    """, unsafe_allow_html=True)
                    
                    if can_edit:
                        with st.expander("✏️ Correct Log Entry", expanded=False):
                            with st.form(f"edit_log_{log_id}"):
                                new_comment_text = st.text_area("Comment", value=log["comment_text"], key=f"ec_{log_id}")
                                new_hours = st.number_input("Hours", min_value=0.0, max_value=24.0, value=float(log["hours_logged"]), step=0.5, key=f"eh_{log_id}")
                                
                                ecol1, ecol2 = st.columns(2)
                                save_btn = ecol1.form_submit_button("💾 Save", type="primary")
                                delete_btn = ecol2.form_submit_button("🗑️ Delete", type="secondary")
                                
                                if save_btn:
                                    if not new_comment_text.strip():
                                        st.error("Comment cannot be empty.")
                                    else:
                                        old_hours = float(log["hours_logged"])
                                        diff = new_hours - old_hours
                                        
                                        exe("UPDATE member_comments SET comment_text = %s, hours_logged = %s WHERE id = %s",
                                            (new_comment_text.strip(), new_hours, log_id))
                                            
                                        if log["task_id"] and diff != 0.0:
                                            exe("UPDATE tasks SET actual_hours = COALESCE(actual_hours, 0) + %s, updated_at = CURRENT_TIMESTAMP WHERE id = %s",
                                                (diff, int(log["task_id"])))
                                                
                                            log_activity(proj_id, member_name, f"updated time log (diff {diff:+.1f}h)", log["task_title"], "actual_hours", f"{old_hours}h", f"{new_hours}h", int(log["task_id"]), log["sprint_id"])
                                        
                                        st.success("Log entry updated successfully!")
                                        st.rerun()
                                        
                                if delete_btn:
                                    old_hours = float(log["hours_logged"])
                                    exe("DELETE FROM member_comments WHERE id = %s", [log_id])
                                    
                                    if log["task_id"] and old_hours > 0.0:
                                        exe("UPDATE tasks SET actual_hours = GREATEST(0.0, COALESCE(actual_hours, 0) - %s), updated_at = CURRENT_TIMESTAMP WHERE id = %s",
                                            (old_hours, int(log["task_id"])))
                                            
                                        log_activity(proj_id, member_name, f"removed progress log ({old_hours}h)", log["task_title"], "actual_hours", f"{old_hours}h", "0.0h", int(log["task_id"]), log["sprint_id"])
                                        
                                    st.success("Log entry deleted.")
                                    st.rerun()

            if total_my_logs > len(my_logs):
                if st.button("Load More Logs", key="load_more_my_logs", use_container_width=True):
                    st.session_state.my_logs_limit += 20
                    st.rerun()

    # ── MANAGER VIEW ──
    elif role == "Manager":
        st.markdown("#### 👥 Team Work Progress Logs")
        
        fcol1, fcol2, fcol3 = st.columns([1, 1, 1.2])
        
        team_df = load_team(proj_id)
        team_opts = {None: "— All Members —"}
        if not team_df.empty:
            team_opts.update({int(row["id"]): row["name"] for _, row in team_df.iterrows()})
            
        sel_member_id = fcol1.selectbox("Filter by Member", list(team_opts.keys()), format_func=lambda x: team_opts[x])
        
        sprint_opts = {None: "— All Sprints —"}
        if not sprints_df.empty:
            sprint_opts.update({int(row["id"]): row["name"] for _, row in sprints_df.iterrows()})
            
        sel_sprint_id = fcol2.selectbox("Filter by Sprint", list(sprint_opts.keys()), format_func=lambda x: sprint_opts[x])
        
        try:
            date_sel = fcol3.date_input("Filter by Date Range", [date.today() - timedelta(days=7), date.today()])
        except:
            date_sel = [date.today() - timedelta(days=7), date.today()]
            
        search_query = st.text_input("🔍 Search comments", "")
        
        if "team_logs_limit" not in st.session_state:
            st.session_state.team_logs_limit = 20

        count_query = """
            SELECT COUNT(*) as cnt
            FROM member_comments mc
            JOIN team_members tm ON mc.member_id = tm.id
            LEFT JOIN sprints s ON mc.sprint_id = s.id
            LEFT JOIN tasks t ON mc.task_id = t.id
            WHERE mc.project_id = %s
        """

        query_str = """
            SELECT mc.*, tm.name as member_name, tm.avatar_color, tm.email as member_email,
                   s.name as sprint_name, t.title as task_title
            FROM member_comments mc
            JOIN team_members tm ON mc.member_id = tm.id
            LEFT JOIN sprints s ON mc.sprint_id = s.id
            LEFT JOIN tasks t ON mc.task_id = t.id
            WHERE mc.project_id = %s
        """
        params = [proj_id]

        if sel_member_id:
            count_query += " AND mc.member_id = %s"
            query_str += " AND mc.member_id = %s"
            params.append(sel_member_id)
        if sel_sprint_id:
            count_query += " AND mc.sprint_id = %s"
            query_str += " AND mc.sprint_id = %s"
            params.append(sel_sprint_id)
        if isinstance(date_sel, (list, tuple)) and len(date_sel) == 2:
            count_query += " AND mc.created_at >= %s AND mc.created_at <= %s"
            query_str += " AND mc.created_at >= %s AND mc.created_at <= %s"
            params.append(str(date_sel[0]) + " 00:00:00")
            params.append(str(date_sel[1]) + " 23:59:59")
        if search_query.strip():
            count_query += " AND mc.comment_text ILIKE %s"
            query_str += " AND mc.comment_text ILIKE %s"
            params.append(f"%{search_query.strip()}%")

        query_str += " ORDER BY mc.id DESC"

        count_df = qry(count_query, params)
        total_team_logs = int(count_df.iloc[0]["cnt"]) if not count_df.empty else 0

        query_str += " LIMIT %s"
        comments_df = qry(query_str, params + [st.session_state.team_logs_limit])
        
        if comments_df.empty:
            st.info("No matching daily progress logs found.")
        else:
            st.caption(f"Showing {len(comments_df)} log entry/entries")
            for _, log in comments_df.iterrows():
                log_id = int(log["id"])
                created_at_dt = to_local_dt(log["created_at"])
                if created_at_dt:
                    time_str = created_at_dt.strftime("%Y-%m-%d %H:%M:%S")
                else:
                    time_str = str(log["created_at"]) if pd.notna(log["created_at"]) else "Unknown Time"
                
                m_name = log["member_name"]
                m_color = log.get("avatar_color", "#3b82f6")
                
                task_info = f" · Task: **{log['task_title']}**" if log["task_title"] else ""
                sprint_info = f" · Sprint: **{log['sprint_name']}**" if log["sprint_name"] else ""
                hours_info = f" · Time Logged: **{log['hours_logged']}h**" if float(log["hours_logged"]) > 0 else ""
                
                st.markdown(f"""
                <div class="card" style="border-left: 3px solid #3b82f6;">
                    <div style="display:flex; gap:10px; align-items:center; margin-bottom:8px;">
                        {av_html(m_name, m_color, 24)}
                        <div>
                            <span style="font-weight:600; color:#e6edf3; font-size:13px;">{m_name}</span>
                            <span style="font-size:11px; color:#8b949e; margin-left:8px;">🕒 {time_str}{sprint_info}{task_info}{hours_info}</span>
                        </div>
                    </div>
                    <p style="color:#e6edf3; font-size:13px; margin:0; padding-left:34px;">{log['comment_text']}</p>
                </div>
                """, unsafe_allow_html=True)


            if total_team_logs > len(comments_df):
                if st.button("Load More Logs", key="load_more_team_logs", use_container_width=True):
                    st.session_state.team_logs_limit += 20
                    st.rerun()                        
    st.markdown('</div>', unsafe_allow_html=True)
